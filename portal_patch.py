
from __future__ import annotations

import base64
import io
import json
import os
import re
import sqlite3
import uuid
from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Any

import fitz
import requests
from flask import abort, flash, g, redirect, request, send_file, url_for
from openpyxl import load_workbook
from PIL import Image, ImageFile, ImageOps, UnidentifiedImageError
from pypdf import PdfReader, PdfWriter
ImageFile.LOAD_TRUNCATED_IMAGES = True


ALLOWED_EXTENSIONS = {".pdf", ".jpg", ".jpeg", ".png"}

MAX_SINGLE_FILE_SIZE = 25 * 1024 * 1024
MAX_FILES_PER_DOCUMENT = 30
MAX_OCR_PAGES = 10

OCR_DOCUMENT_TYPES = {
    "parent_passport",
    "child_passport",
    "parent_snils",
    "child_snils",
    "birth_certificate",
}

DATE_COLUMNS = (3, 10, 11, 12)


def apply_patch(app, namespace: dict[str, Any]) -> None:
    base_dir: Path = namespace["BASE_DIR"]
    database_path: Path = namespace["DATABASE_PATH"]
    upload_dir: Path = namespace["UPLOAD_DIR"]

    get_db = namespace["get_db"]
    get_student_or_404 = namespace["get_student_or_404"]
    roles_required = namespace["roles_required"]
    audit = namespace["audit"]
    render_page = namespace["render_page"]
    calculate_age = namespace["calculate_age"]
    full_name = namespace["full_name"]

    document_types = namespace["DOCUMENT_TYPES"]

    app.config["MAX_CONTENT_LENGTH"] = 150 * 1024 * 1024

    template_path = base_dir / "export_template.xlsx"

    def migrate_database() -> None:
        if not database_path.exists():
            return

        connection = sqlite3.connect(database_path)

        try:
            columns = {
                row[1]
                for row in connection.execute(
                    "PRAGMA table_info(documents)"
                ).fetchall()
            }

            migrations = {
                "page_count": (
                    "ALTER TABLE documents "
                    "ADD COLUMN page_count INTEGER NOT NULL DEFAULT 1"
                ),
                "ocr_text": (
                    "ALTER TABLE documents "
                    "ADD COLUMN ocr_text TEXT"
                ),
                "ocr_json": (
                    "ALTER TABLE documents "
                    "ADD COLUMN ocr_json TEXT"
                ),
            }

            for column_name, sql in migrations.items():
                if column_name not in columns:
                    connection.execute(sql)

            connection.commit()

        finally:
            connection.close()

    original_init_db = namespace["init_db"]

    def init_db_v2() -> None:
        original_init_db()
        migrate_database()

    namespace["init_db"] = init_db_v2

    if database_path.exists():
        migrate_database()

    def clean_original_name(filename: str) -> str:
        filename = str(filename or "").replace("\\", "/")
        filename = Path(filename).name.strip()

        return filename or "document"

    def extension_from_name(filename: str) -> str:
        return Path(clean_original_name(filename)).suffix.lower()

    def validate_upload(upload) -> tuple[str, bytes]:
        original_name = clean_original_name(upload.filename)
        extension = extension_from_name(original_name)

        if extension not in ALLOWED_EXTENSIONS:
            raise ValueError(
                f"Файл «{original_name}»: разрешены только "
                "PDF, JPG, JPEG и PNG."
            )

        content = upload.read()

        if not content:
            raise ValueError(
                f"Файл «{original_name}» пустой."
            )

        if len(content) > MAX_SINGLE_FILE_SIZE:
            raise ValueError(
                f"Файл «{original_name}» больше 25 МБ."
            )

        pdf_position = content[:1024].find(b"%PDF")

        if extension == ".pdf" or pdf_position >= 0:
            if pdf_position < 0:
                raise ValueError(
                    f"Файл «{original_name}» имеет расширение PDF, "
                    "но не является корректным PDF."
                )

            pdf_content = content[pdf_position:]

            try:
                reader = PdfReader(
                    io.BytesIO(pdf_content)
                )

                if reader.is_encrypted:
                    raise ValueError(
                        f"Файл «{original_name}» защищен паролем."
                    )

                if not reader.pages:
                    raise ValueError(
                        f"В файле «{original_name}» нет страниц."
                    )

            except ValueError:
                raise

            except Exception as error:
                raise ValueError(
                    f"Не удалось прочитать PDF "
                    f"«{original_name}»: {error}"
                ) from error

            return original_name, pdf_content

        try:
            image = Image.open(
                io.BytesIO(content)
            )

            actual_format = str(
                image.format or ""
            ).upper()

            if actual_format not in {
                "JPEG",
                "JPG",
                "PNG",
            }:
                raise ValueError(
                    f"Файл «{original_name}» фактически имеет "
                    f"формат {actual_format or 'неизвестный'}, "
                    "а разрешены только JPEG и PNG."
                )

            image.load()

            if image.width < 1 or image.height < 1:
                raise ValueError(
                    f"В изображении «{original_name}» "
                    "нет корректной страницы."
                )

        except ValueError:
            raise

        except (
            UnidentifiedImageError,
            OSError,
            SyntaxError,
        ) as error:
            raise ValueError(
                f"Не удалось прочитать изображение "
                f"«{original_name}». "
                "Откройте его на компьютере и сохраните "
                "повторно как JPG или PNG. "
                f"Техническая причина: {error}"
            ) from error

        return original_name, content

    def image_to_pdf_reader(content: bytes) -> PdfReader:
        image = Image.open(io.BytesIO(content))
        image = ImageOps.exif_transpose(image)

        if image.mode not in ("RGB", "L"):
            image = image.convert("RGB")
        elif image.mode == "L":
            image = image.convert("RGB")

        output = io.BytesIO()

        image.save(
            output,
            format="PDF",
            resolution=150.0,
        )

        output.seek(0)

        return PdfReader(output)

    def merge_uploaded_files(uploads) -> tuple[bytes, list[str], int]:
        if len(uploads) > MAX_FILES_PER_DOCUMENT:
            raise ValueError(
                f"За один раз можно загрузить не более "
                f"{MAX_FILES_PER_DOCUMENT} файлов."
            )

        writer = PdfWriter()
        original_names: list[str] = []
        page_count = 0

        for upload in uploads:
            original_name, content = validate_upload(upload)
            extension = extension_from_name(original_name)

            original_names.append(original_name)

            if extension == ".pdf":
                reader = PdfReader(io.BytesIO(content))
            else:
                reader = image_to_pdf_reader(content)

            for page in reader.pages:
                writer.add_page(page)
                page_count += 1

        if page_count == 0:
            raise ValueError("В выбранных файлах нет страниц.")

        result = io.BytesIO()
        writer.write(result)

        return result.getvalue(), original_names, page_count

    def normalize_gender(value: str | None) -> str:
        normalized = str(value or "").strip().lower()

        if normalized in {
            "жен",
            "жен.",
            "женский",
            "женщина",
            "female",
            "f",
        }:
            return "жен"

        if normalized in {
            "муж",
            "муж.",
            "мужской",
            "мужчина",
            "male",
            "m",
        }:
            return "муж"

        return ""

    def iso_date(value: str | None):
        if not value:
            return ""

        value = str(value).strip()

        for date_format in (
            "%Y-%m-%d",
            "%d.%m.%Y",
            "%d/%m/%Y",
            "%d-%m-%Y",
        ):
            try:
                return datetime.strptime(
                    value,
                    date_format,
                ).date()

            except ValueError:
                continue

        return ""

    def format_snils(value: str) -> str:
        digits = re.sub(r"\D", "", value or "")

        if len(digits) != 11:
            return value.strip()

        return (
            f"{digits[0:3]}-{digits[3:6]}-"
            f"{digits[6:9]} {digits[9:11]}"
        )

    def smart_title(value: str) -> str:
        return " ".join(
            part[:1].upper() + part[1:].lower()
            for part in str(value or "").strip().split()
        )

    def ocr_is_configured() -> bool:
        return bool(
            os.getenv("YANDEX_FOLDER_ID")
            and (
                os.getenv("YANDEX_VISION_API_KEY")
                or os.getenv("YANDEX_VISION_IAM_TOKEN")
            )
        )

    def ocr_headers() -> dict[str, str]:
        api_key = os.getenv("YANDEX_VISION_API_KEY")
        iam_token = os.getenv("YANDEX_VISION_IAM_TOKEN")
        folder_id = os.getenv("YANDEX_FOLDER_ID", "")

        if api_key:
            authorization = f"Api-Key {api_key}"
        elif iam_token:
            authorization = f"Bearer {iam_token}"
        else:
            raise RuntimeError(
                "Не задан YANDEX_VISION_API_KEY "
                "или YANDEX_VISION_IAM_TOKEN."
            )

        return {
            "Content-Type": "application/json",
            "Authorization": authorization,
            "x-folder-id": folder_id,
            "x-data-logging-enabled": "false",
        }

    def pdf_pages_to_jpeg(
        document_path: Path,
    ) -> list[bytes]:
        result: list[bytes] = []

        document = fitz.open(document_path)

        try:
            page_limit = min(
                len(document),
                MAX_OCR_PAGES,
            )

            for page_number in range(page_limit):
                page = document.load_page(page_number)

                pixmap = page.get_pixmap(
                    matrix=fitz.Matrix(1.7, 1.7),
                    alpha=False,
                )

                jpeg_data = pixmap.tobytes("jpeg")

                if len(jpeg_data) > 9 * 1024 * 1024:
                    image = Image.open(io.BytesIO(jpeg_data))
                    image.thumbnail((2500, 2500))

                    compressed = io.BytesIO()

                    image.save(
                        compressed,
                        format="JPEG",
                        quality=80,
                        optimize=True,
                    )

                    jpeg_data = compressed.getvalue()

                result.append(jpeg_data)

        finally:
            document.close()

        return result

    def find_entities(value: Any) -> list[dict[str, str]]:
        if isinstance(value, dict):
            entities = value.get("entities")

            if isinstance(entities, list):
                return [
                    item
                    for item in entities
                    if isinstance(item, dict)
                ]

            for child in value.values():
                found = find_entities(child)

                if found:
                    return found

        if isinstance(value, list):
            for child in value:
                found = find_entities(child)

                if found:
                    return found

        return []

    def collect_text(value: Any) -> list[str]:
        texts: list[str] = []

        if isinstance(value, dict):
            for key, child in value.items():
                if (
                    key in {"fullText", "text"}
                    and isinstance(child, str)
                    and child.strip()
                ):
                    texts.append(child.strip())
                else:
                    texts.extend(collect_text(child))

        elif isinstance(value, list):
            for child in value:
                texts.extend(collect_text(child))

        return texts

    def recognize_document(
        document_path: Path,
        document_type: str,
    ) -> dict[str, Any]:
        if not ocr_is_configured():
            raise RuntimeError(
                "OCR не настроен. Необходимо задать "
                "YANDEX_FOLDER_ID и YANDEX_VISION_API_KEY."
            )

        model = (
            "passport"
            if document_type in {
                "parent_passport",
                "child_passport",
            }
            else "page"
        )

        endpoint = (
            "https://ocr.api.cloud.yandex.net/"
            "ocr/v1/recognizeText"
        )

        all_entities: dict[str, str] = {}
        all_text: list[str] = []
        raw_responses: list[dict[str, Any]] = []

        for jpeg_data in pdf_pages_to_jpeg(document_path):
            payload = {
                "mimeType": "JPEG",
                "languageCodes": ["ru", "en"],
                "model": model,
                "content": base64.b64encode(
                    jpeg_data
                ).decode("ascii"),
            }

            response = requests.post(
                endpoint,
                headers=ocr_headers(),
                json=payload,
                timeout=90,
            )

            response.raise_for_status()
            response_data = response.json()

            raw_responses.append(response_data)

            for entity in find_entities(response_data):
                name = str(entity.get("name", "")).strip()
                text = str(entity.get("text", "")).strip()

                if name and text and name not in all_entities:
                    all_entities[name] = text

            all_text.extend(collect_text(response_data))

            if (
                model == "passport"
                and all_entities.get("surname")
                and all_entities.get("number")
            ):
                break

        return {
            "entities": all_entities,
            "text": "\n".join(dict.fromkeys(all_text)),
            "responses": raw_responses,
        }

    def date_from_ocr(value: str) -> str:
        match = re.search(
            r"\b(\d{2})[./-](\d{2})[./-](\d{4})\b",
            value or "",
        )

        if not match:
            return ""

        day, month, year = match.groups()

        try:
            parsed = datetime.strptime(
                f"{day}.{month}.{year}",
                "%d.%m.%Y",
            )

        except ValueError:
            return ""

        return parsed.strftime("%Y-%m-%d")

    def extract_updates(
        document_type: str,
        recognition: dict[str, Any],
    ) -> dict[str, str]:
        entities = {
            str(key).lower(): str(value).strip()
            for key, value in recognition.get(
                "entities",
                {},
            ).items()
        }

        text = str(recognition.get("text", ""))
        updates: dict[str, str] = {}

        if document_type in {
            "parent_passport",
            "child_passport",
        }:
            number = re.sub(
                r"\D",
                "",
                entities.get("number", ""),
            )

            passport_series = (
                number[:4]
                if len(number) >= 10
                else ""
            )

            passport_number = (
                number[-6:]
                if len(number) >= 10
                else ""
            )

            gender = normalize_gender(
                entities.get("gender", "")
            )

            birth_date = date_from_ocr(
                entities.get("birth_date", "")
            )

            surname = smart_title(
                entities.get("surname", "")
            )

            first_name = smart_title(
                entities.get("name", "")
            )

            middle_name = smart_title(
                entities.get("middle_name", "")
            )

            if document_type == "parent_passport":
                updates.update(
                    {
                        "parent_last_name": surname,
                        "parent_first_name": first_name,
                        "parent_middle_name": middle_name,
                        "parent_gender": gender,
                        "parent_birth_date": birth_date,
                        "parent_passport_series": passport_series,
                        "parent_passport_number": passport_number,
                    }
                )

            else:
                updates.update(
                    {
                        "last_name": surname,
                        "first_name": first_name,
                        "middle_name": middle_name,
                        "gender": gender,
                        "birth_date": birth_date,
                        "child_passport_series": passport_series,
                        "child_passport_number": passport_number,
                    }
                )

        elif document_type in {
            "parent_snils",
            "child_snils",
        }:
            match = re.search(
                r"\b\d{3}[-\s]?\d{3}[-\s]?"
                r"\d{3}\s?\d{2}\b",
                text,
            )

            if match:
                snils = format_snils(match.group(0))

                target_field = (
                    "parent_snils"
                    if document_type == "parent_snils"
                    else "student_snils"
                )

                updates[target_field] = snils

        elif document_type == "birth_certificate":
            certificate_match = re.search(
                r"\b([IVXLCDM]+[\s\-–—]*[А-ЯЁ]{2})"
                r"\s*(?:№|N)?\s*(\d{6})\b",
                text.upper(),
            )

            if certificate_match:
                updates["birth_certificate_series"] = (
                    certificate_match.group(1)
                    .replace(" ", "")
                    .replace("–", "-")
                    .replace("—", "-")
                )

                updates["birth_certificate_number"] = (
                    certificate_match.group(2)
                )

            dates = re.findall(
                r"\b\d{2}[./-]\d{2}[./-]\d{4}\b",
                text,
            )

            if dates:
                updates["birth_date"] = date_from_ocr(
                    dates[0]
                )

        return {
            key: value
            for key, value in updates.items()
            if value
        }

    def apply_empty_student_fields(
        student_id: int,
        updates: dict[str, str],
    ) -> tuple[list[str], list[str]]:
        student = get_db().execute(
            "SELECT * FROM students WHERE id = ?",
            (student_id,),
        ).fetchone()

        if not student:
            raise RuntimeError("Карточка ученика не найдена.")

        fields_to_update: dict[str, str] = {}
        mismatches: list[str] = []

        for field_name, new_value in updates.items():
            current_value = str(
                student[field_name] or ""
            ).strip()

            if not current_value:
                fields_to_update[field_name] = new_value

            elif (
                re.sub(r"\s+", "", current_value.lower())
                != re.sub(r"\s+", "", new_value.lower())
            ):
                mismatches.append(field_name)

        if fields_to_update:
            assignments = ", ".join(
                f"{field_name} = ?"
                for field_name in fields_to_update
            )

            values = list(fields_to_update.values())

            get_db().execute(
                f"""
                UPDATE students
                SET {assignments},
                    updated_at = ?
                WHERE id = ?
                """,
                [
                    *values,
                    datetime.now().isoformat(
                        timespec="seconds"
                    ),
                    student_id,
                ],
            )

            get_db().commit()

        return list(fields_to_update), mismatches

    def process_document_ocr(
        document_id: int,
    ) -> tuple[list[str], list[str]]:
        document = get_db().execute(
            """
            SELECT *
            FROM documents
            WHERE id = ?
            """,
            (document_id,),
        ).fetchone()

        if not document:
            raise RuntimeError("Документ не найден.")

        if document["document_type"] not in OCR_DOCUMENT_TYPES:
            return [], []

        document_path = upload_dir / document["stored_name"]

        get_db().execute(
            """
            UPDATE documents
            SET ocr_status = 'processing'
            WHERE id = ?
            """,
            (document_id,),
        )

        get_db().commit()

        try:
            recognition = recognize_document(
                document_path,
                document["document_type"],
            )

            updates = extract_updates(
                document["document_type"],
                recognition,
            )

            filled_fields, mismatches = (
                apply_empty_student_fields(
                    document["student_id"],
                    updates,
                )
            )

            get_db().execute(
                """
                UPDATE documents
                SET
                    ocr_status = 'done',
                    ocr_text = ?,
                    ocr_json = ?
                WHERE id = ?
                """,
                (
                    recognition.get("text", ""),
                    json.dumps(
                        recognition,
                        ensure_ascii=False,
                    ),
                    document_id,
                ),
            )

            get_db().commit()

            return filled_fields, mismatches

        except Exception as error:
            get_db().execute(
                """
                UPDATE documents
                SET
                    ocr_status = 'error',
                    ocr_text = ?
                WHERE id = ?
                """,
                (
                    str(error),
                    document_id,
                ),
            )

            get_db().commit()

            raise

    def upload_document_v2(
        student_id: int,
        document_type: str,
    ):
        student = get_student_or_404(student_id)

        if student["status"] not in (
            "draft",
            "correction",
        ):
            abort(
                400,
                "Карточка заблокирована для редактирования.",
            )

        if document_type not in document_types:
            abort(404)

        existing_documents = get_db().execute(
            """
            SELECT *
            FROM documents
            WHERE student_id = ?
            ORDER BY uploaded_at DESC
            """,
            (student_id,),
        ).fetchall()

        if request.method == "POST":
            reuse_document_id = request.form.get(
                "reuse_document_id",
                "",
            ).strip()

            original_name = ""
            stored_name = ""
            mime_type = "application/pdf"
            source_document_id = None
            page_count = 1

            if reuse_document_id:
                source = get_db().execute(
                    """
                    SELECT *
                    FROM documents
                    WHERE id = ?
                      AND student_id = ?
                    """,
                    (
                        int(reuse_document_id),
                        student_id,
                    ),
                ).fetchone()

                if not source:
                    abort(
                        400,
                        "Исходный документ не найден.",
                    )

                original_name = source["original_name"]
                stored_name = source["stored_name"]
                mime_type = source["mime_type"]
                source_document_id = source["id"]

                try:
                    page_count = int(
                        source["page_count"] or 1
                    )
                except (KeyError, TypeError, ValueError):
                    page_count = 1

            else:
                uploads = [
                    upload
                    for upload in request.files.getlist(
                        "files"
                    )
                    if upload and upload.filename
                ]

                if not uploads:
                    flash(
                        "Выберите один или несколько файлов.",
                        "error",
                    )

                    return redirect(request.url)

                try:
                    (
                        merged_pdf,
                        original_names,
                        page_count,
                    ) = merge_uploaded_files(uploads)

                except ValueError as error:
                    flash(str(error), "error")
                    return redirect(request.url)

                stored_name = f"{uuid.uuid4().hex}.pdf"

                (upload_dir / stored_name).write_bytes(
                    merged_pdf
                )

                original_name = " + ".join(
                    original_names
                )[:1500]

            latest_version = get_db().execute(
                """
                SELECT COALESCE(MAX(version), 0)
                FROM documents
                WHERE student_id = ?
                  AND document_type = ?
                """,
                (
                    student_id,
                    document_type,
                ),
            ).fetchone()[0]

            cursor = get_db().execute(
                """
                INSERT INTO documents (
                    student_id,
                    document_type,
                    original_name,
                    stored_name,
                    mime_type,
                    version,
                    source_document_id,
                    uploaded_by,
                    uploaded_at,
                    ocr_status,
                    page_count
                )
                VALUES (
                    ?, ?, ?, ?, ?, ?, ?, ?, ?,
                    'not_configured', ?
                )
                """,
                (
                    student_id,
                    document_type,
                    original_name,
                    stored_name,
                    mime_type,
                    latest_version + 1,
                    source_document_id,
                    g.current_user["id"],
                    datetime.now().isoformat(
                        timespec="seconds"
                    ),
                    page_count,
                ),
            )

            document_id = cursor.lastrowid

            get_db().execute(
                """
                UPDATE students
                SET updated_at = ?
                WHERE id = ?
                """,
                (
                    datetime.now().isoformat(
                        timespec="seconds"
                    ),
                    student_id,
                ),
            )

            get_db().commit()

            audit(
                "document_uploaded",
                student_id,
                (
                    f"{document_type}; "
                    f"версия {latest_version + 1}; "
                    f"страниц {page_count}"
                ),
            )

            message = (
                f"Документ загружен. "
                f"Страниц: {page_count}."
            )

            if document_type in OCR_DOCUMENT_TYPES:
                if ocr_is_configured():
                    try:
                        filled, mismatches = (
                            process_document_ocr(
                                document_id
                            )
                        )

                        if filled:
                            message += (
                                " Автоматически заполнены поля: "
                                + ", ".join(filled)
                                + "."
                            )

                        if mismatches:
                            message += (
                                " Найдены расхождения с ранее "
                                "введенными полями: "
                                + ", ".join(mismatches)
                                + "."
                            )

                    except Exception as error:
                        message += (
                            " Документ сохранен, но OCR "
                            f"завершился ошибкой: {error}"
                        )

                else:
                    message += (
                        " OCR пока не настроен — файл "
                        "сохранен без распознавания."
                    )

            flash(message, "success")

            return redirect(
                url_for(
                    "student_detail",
                    student_id=student_id,
                )
            )

        body = """
        <h1>{{ document_config.name }}</h1>

        <div class="alert info">
            Можно выбрать несколько PDF, JPG, JPEG или PNG.
            Файлы будут объединены в один многостраничный PDF
            в том порядке, в котором их передал браузер.
            Старые версии не удаляются.
        </div>

        <form
            class="form-section"
            method="post"
            enctype="multipart/form-data"
        >
            <input
                type="hidden"
                name="csrf_token"
                value="{{ csrf_token }}"
            >

            <div>
                <label>
                    Новый файл или несколько страниц
                </label>

                <input
                    type="file"
                    name="files"
                    accept=".pdf,.jpg,.jpeg,.png"
                    multiple
                >

                <p class="muted">
                    До 30 файлов за один раз.
                    Каждый файл — не более 25 МБ.
                </p>
            </div>

            {% if existing_documents %}
                <h3>
                    Или использовать ранее загруженный файл
                </h3>

                <select name="reuse_document_id">
                    <option value="">
                        Не использовать
                    </option>

                    {% for document in existing_documents %}
                        <option value="{{ document.id }}">
                            {{
                                document_types[
                                    document.document_type
                                ].name
                            }}
                            — {{ document.original_name }}
                            — версия {{ document.version }}
                        </option>
                    {% endfor %}
                </select>
            {% endif %}

            <div
                class="space"
                style="
                    display:flex;
                    gap:12px;
                    flex-wrap:wrap;
                    align-items:center;
                "
            >
                <button
                    class="btn btn-primary"
                    type="submit"
                >
                    Сохранить документ
                </button>

                <a
                    class="btn btn-secondary"
                    href="{{ url_for(
                        'student_detail',
                        student_id=student.id
                    ) }}"
                >
                    Назад
                </a>
            </div>
        </form>
        """

        return render_page(
            "Загрузка документа",
            body,
            student=student,
            document_config=document_types[
                document_type
            ],
            existing_documents=existing_documents,
            document_types=document_types,
        )

    def find_template() -> Path | None:
        if template_path.exists():
            return template_path

        patterns = (
            "*писок*обучающихся*законных*"
            "представителей*.xlsx",
            "*обучающихся*представителей*.xlsx",
        )

        for pattern in patterns:
            for candidate in base_dir.glob(pattern):
                if not candidate.name.startswith("~$"):
                    return candidate

        return None

    def copy_row_style(
        worksheet,
        source_styles,
        target_row: int,
    ) -> None:
        for column_index, style_data in enumerate(
            source_styles,
            start=1,
        ):
            target_cell = worksheet.cell(
                row=target_row,
                column=column_index,
            )

            (
                style,
                number_format,
                alignment,
                font,
                fill,
                border,
                protection,
            ) = style_data

            target_cell._style = copy(style)
            target_cell.number_format = number_format
            target_cell.alignment = copy(alignment)
            target_cell.font = copy(font)
            target_cell.fill = copy(fill)
            target_cell.border = copy(border)
            target_cell.protection = copy(
                protection
            )

    def export_xlsx_v2():
        branch_id = request.args.get(
            "branch_id",
            type=int,
        )

        class_number = request.args.get(
            "class_number",
            type=int,
        )

        if not branch_id or not class_number:
            abort(400)

        branch = get_db().execute(
            "SELECT * FROM branches WHERE id = ?",
            (branch_id,),
        ).fetchone()

        if not branch:
            abort(404)

        export_template = find_template()

        if not export_template:
            abort(
                500,
                "Не найден файл export_template.xlsx "
                "в папке портала.",
            )

        students_rows = get_db().execute(
            """
            SELECT *
            FROM students
            WHERE branch_id = ?
              AND class_number = ?
              AND status = 'ready'
            ORDER BY last_name, first_name
            """,
            (
                branch_id,
                class_number,
            ),
        ).fetchall()

        workbook = load_workbook(export_template)

        worksheet = (
            workbook["Импорт"]
            if "Импорт" in workbook.sheetnames
            else workbook.active
        )

        style_source_row = (
            2
            if worksheet.max_row >= 2
            else 1
        )

        source_styles = []

        for column_index in range(1, 22):
            source_cell = worksheet.cell(
                row=style_source_row,
                column=column_index,
            )

            source_styles.append(
                (
                    copy(source_cell._style),
                    source_cell.number_format,
                    copy(source_cell.alignment),
                    copy(source_cell.font),
                    copy(source_cell.fill),
                    copy(source_cell.border),
                    copy(source_cell.protection),
                )
            )

        if worksheet.max_row >= 2:
            worksheet.delete_rows(
                2,
                worksheet.max_row - 1,
            )

        for student in students_rows:
            age = calculate_age(
                student["birth_date"]
            )

            birth_series = (
                student["birth_certificate_series"]
                if age < 14
                else ""
            )

            birth_number = (
                student["birth_certificate_number"]
                if age < 14
                else ""
            )

            passport_series = (
                student["child_passport_series"]
                if age >= 14
                else ""
            )

            passport_number = (
                student["child_passport_number"]
                if age >= 14
                else ""
            )

            row_values = [
                full_name(
                    student["parent_last_name"],
                    student["parent_first_name"],
                    student["parent_middle_name"],
                ),
                normalize_gender(
                    student["parent_gender"]
                ),
                iso_date(
                    student["parent_birth_date"]
                ),
                student["parent_snils"] or "",
                student[
                    "parent_passport_series"
                ] or "",
                student[
                    "parent_passport_number"
                ] or "",
                student["relation_type"] or "",
                full_name(
                    student["last_name"],
                    student["first_name"],
                    student["middle_name"],
                ),
                normalize_gender(
                    student["gender"]
                ),
                iso_date(student["birth_date"]),
                iso_date(
                    student["arrival_school_date"]
                ),
                iso_date(
                    student["arrival_class_date"]
                ),
                student["student_snils"] or "",
                birth_series,
                birth_number,
                passport_series,
                passport_number,
                student["parent_phone"] or "",
                student["parent_email"] or "",
                student["student_phone"] or "",
                student["student_email"] or "",
            ]

            worksheet.append(row_values)
            target_row = worksheet.max_row

            copy_row_style(
                worksheet,
                source_styles,
                target_row,
            )

            for date_column in DATE_COLUMNS:
                date_cell = worksheet.cell(
                    row=target_row,
                    column=date_column,
                )

                date_cell.number_format = "dd.mm.yyyy"

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        audit(
            "xlsx_exported",
            details=(
                f"branch={branch['name']}; "
                f"class={class_number}; "
                f"rows={len(students_rows)}; "
                "template=export_template.xlsx"
            ),
        )

        safe_branch_name = re.sub(
            r'[\\/:*?"<>|]+',
            "_",
            branch["name"],
        )

        filename = (
            f"Дневник_{safe_branch_name}_"
            f"{class_number}_класс.xlsx"
        )

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype=(
                "application/vnd.openxmlformats-"
                "officedocument.spreadsheetml.sheet"
            ),
        )

    app.view_functions["upload_document"] = (
        roles_required("branch")(
            upload_document_v2
        )
    )

    app.view_functions["export_xlsx"] = (
        roles_required(
            "attestation",
            "admin",
        )(
            export_xlsx_v2
        )
    )
