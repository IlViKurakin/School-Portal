from __future__ import annotations

import hmac
import io
import os
import secrets
import sqlite3
import uuid
from datetime import date, datetime
from functools import wraps
from pathlib import Path
from typing import Any

from flask import (
    Flask,
    abort,
    flash,
    g,
    redirect,
    render_template_string,
    request,
    send_file,
    send_from_directory,
    session,
    url_for,
)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename


# ============================================================
# Конфигурация
# ============================================================

BASE_DIR = Path(__file__).resolve().parent
DATABASE_PATH = BASE_DIR / "portal.db"
UPLOAD_DIR = BASE_DIR / "uploads"

ALLOWED_EXTENSIONS = {"pdf", "jpg", "jpeg", "png"}
MAX_FILE_SIZE = 25 * 1024 * 1024  # 25 МБ

app = Flask(__name__)
app.config.update(
    SECRET_KEY=os.getenv("PORTAL_SECRET_KEY", secrets.token_hex(32)),
    MAX_CONTENT_LENGTH=MAX_FILE_SIZE,
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
)

if os.getenv("PORTAL_HTTPS", "0") == "1":
    app.config["SESSION_COOKIE_SECURE"] = True

UPLOAD_DIR.mkdir(exist_ok=True)


# ============================================================
# Справочники
# ============================================================

STATUS_LABELS = {
    "draft": "Черновик",
    "submitted": "Передано в отдел аттестации",
    "in_review": "На ручной проверке",
    "correction": "Требует исправления",
    "ready": "Готово к зачислению",
    "enrolled": "Зачислен",
}

STATUS_CLASSES = {
    "draft": "gray",
    "submitted": "yellow",
    "in_review": "blue",
    "correction": "red",
    "ready": "light-green",
    "enrolled": "green",
}

DOCUMENT_TYPES = {
    "attachment_application": {
        "name": "Заявление на прикрепление",
        "group": "Заявления и согласия",
        "rule": "all",
    },
    "withdrawal_application": {
        "name": "Заявление на отчисление",
        "group": "Заявления и согласия",
        "rule": "optional",
    },
    "birth_certificate": {
        "name": "Свидетельство о рождении ребенка",
        "group": "Документы ребенка",
        "rule": "all",
    },
    "child_registration": {
        "name": "Подтверждение регистрации ребенка",
        "group": "Документы ребенка",
        "rule": "all",
        "help": (
            "Допускается свидетельство о регистрации по месту жительства, "
            "по месту пребывания, страница паспорта с регистрацией или иной "
            "официальный подтверждающий документ."
        ),
    },
    "child_passport": {
        "name": "Паспорт ребенка",
        "group": "Документы ребенка",
        "rule": "age_14",
    },
    "parent_passport": {
        "name": "Паспорт законного представителя",
        "group": "Документы законного представителя",
        "rule": "all",
    },
    "child_snils": {
        "name": "СНИЛС ребенка",
        "group": "Документы ребенка",
        "rule": "all",
    },
    "parent_snils": {
        "name": "СНИЛС законного представителя",
        "group": "Документы законного представителя",
        "rule": "all",
    },
    "personal_file": {
        "name": "Личное дело или справки предыдущих аттестаций",
        "group": "Документы ребенка",
        "rule": "class_2",
    },
    "parent_consent": {
        "name": "Согласие законного представителя",
        "group": "Заявления и согласия",
        "rule": "all",
    },
    "child_consent": {
        "name": "Согласие ребенка",
        "group": "Заявления и согласия",
        "rule": "age_14",
    },
    "education_notice": {
        "name": "Уведомление в орган образования",
        "group": "Заявления и согласия",
        "rule": "all",
    },
    "grade9_certificate": {
        "name": "Аттестат за 9 класс",
        "group": "Документы ребенка",
        "rule": "class_10",
    },
    "oge_results": {
        "name": "Результаты ОГЭ",
        "group": "Документы ребенка",
        "rule": "class_10",
    },
    "relation_proof": {
        "name": "Документ о смене фамилии, опеке или усыновлении",
        "group": "Документы законного представителя",
        "rule": "relation_proof",
    },
    "citizenship_mark": {
        "name": "Отметка о российском гражданстве",
        "group": "Документы ребенка",
        "rule": "russian_citizen",
        "help": "Оборотная сторона свидетельства о рождении.",
    },
}

COMMENT_CATEGORIES = [
    "Документ отсутствует",
    "Загружен неверный документ",
    "Плохое качество изображения",
    "Отсутствует обязательная страница",
    "Не совпадает ФИО",
    "Не совпадает дата рождения",
    "Не совпадает номер документа",
    "Отсутствует подпись",
    "Отсутствует печать или штамп",
    "Нужен подтверждающий документ",
    "Другое",
]


# ============================================================
# База данных
# ============================================================

def get_db() -> sqlite3.Connection:
    if "db" not in g:
        connection = sqlite3.connect(DATABASE_PATH)
        connection.row_factory = sqlite3.Row
        connection.execute("PRAGMA foreign_keys = ON")
        g.db = connection

    return g.db


@app.teardown_appcontext
def close_db(_: Any = None) -> None:
    db = g.pop("db", None)
    if db is not None:
        db.close()


def init_db() -> None:
    db = sqlite3.connect(DATABASE_PATH)
    db.execute("PRAGMA foreign_keys = ON")

    db.executescript(
        """
        CREATE TABLE IF NOT EXISTS branches (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE
        );

        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            email TEXT NOT NULL UNIQUE,
            password_hash TEXT NOT NULL,
            full_name TEXT NOT NULL,
            role TEXT NOT NULL CHECK (
                role IN ('branch', 'attestation', 'admin')
            ),
            branch_id INTEGER,
            active INTEGER NOT NULL DEFAULT 1,
            created_at TEXT NOT NULL,
            FOREIGN KEY (branch_id) REFERENCES branches(id)
        );

        CREATE TABLE IF NOT EXISTS students (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            branch_id INTEGER NOT NULL,
            created_by INTEGER NOT NULL,
            assigned_to INTEGER,

            last_name TEXT NOT NULL,
            first_name TEXT NOT NULL,
            middle_name TEXT,
            gender TEXT,
            birth_date TEXT NOT NULL,
            citizenship TEXT NOT NULL DEFAULT 'РФ',
            class_number INTEGER NOT NULL,

            arrival_school_date TEXT,
            arrival_class_date TEXT,

            student_snils TEXT,
            birth_certificate_series TEXT,
            birth_certificate_number TEXT,
            child_passport_series TEXT,
            child_passport_number TEXT,
            student_phone TEXT,
            student_email TEXT,

            parent_last_name TEXT NOT NULL,
            parent_first_name TEXT NOT NULL,
            parent_middle_name TEXT,
            parent_gender TEXT,
            parent_birth_date TEXT,
            parent_snils TEXT,
            parent_passport_series TEXT,
            parent_passport_number TEXT,
            relation_type TEXT NOT NULL,
            parent_phone TEXT,
            parent_email TEXT,

            needs_relation_proof INTEGER NOT NULL DEFAULT 0,
            status TEXT NOT NULL DEFAULT 'draft',

            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,

            FOREIGN KEY (branch_id) REFERENCES branches(id),
            FOREIGN KEY (created_by) REFERENCES users(id),
            FOREIGN KEY (assigned_to) REFERENCES users(id)
        );

        CREATE TABLE IF NOT EXISTS documents (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            student_id INTEGER NOT NULL,
            document_type TEXT NOT NULL,
            original_name TEXT NOT NULL,
            stored_name TEXT NOT NULL,
            mime_type TEXT,
            version INTEGER NOT NULL,
            source_document_id INTEGER,
            uploaded_by INTEGER NOT NULL,
            uploaded_at TEXT NOT NULL,
            ocr_status TEXT NOT NULL DEFAULT 'not_configured',

            FOREIGN KEY (student_id) REFERENCES students(id),
            FOREIGN KEY (uploaded_by) REFERENCES users(id),
            FOREIGN KEY (source_document_id) REFERENCES documents(id)
        );

        CREATE TABLE IF NOT EXISTS comments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            student_id INTEGER NOT NULL,
            document_type TEXT,
            category TEXT NOT NULL,
            text TEXT NOT NULL,
            is_open INTEGER NOT NULL DEFAULT 1,
            created_by INTEGER NOT NULL,
            created_at TEXT NOT NULL,
            resolved_by INTEGER,
            resolved_at TEXT,

            FOREIGN KEY (student_id) REFERENCES students(id),
            FOREIGN KEY (created_by) REFERENCES users(id),
            FOREIGN KEY (resolved_by) REFERENCES users(id)
        );

        CREATE TABLE IF NOT EXISTS audit_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            student_id INTEGER,
            action TEXT NOT NULL,
            details TEXT,
            created_at TEXT NOT NULL,

            FOREIGN KEY (user_id) REFERENCES users(id),
            FOREIGN KEY (student_id) REFERENCES students(id)
        );

        CREATE INDEX IF NOT EXISTS idx_students_branch
            ON students(branch_id);

        CREATE INDEX IF NOT EXISTS idx_students_status
            ON students(status);

        CREATE INDEX IF NOT EXISTS idx_documents_student
            ON documents(student_id);

        CREATE INDEX IF NOT EXISTS idx_comments_student
            ON comments(student_id);
        """
    )

    now = datetime.now().isoformat(timespec="seconds")

    for branch_name in ("Казань", "Новороссийск"):
        db.execute(
            "INSERT OR IGNORE INTO branches(name) VALUES (?)",
            (branch_name,),
        )

    db.commit()

    branches = {
        row[1]: row[0]
        for row in db.execute("SELECT id, name FROM branches").fetchall()
    }

    demo_users = [
        (
            "admin@top-academy.ru",
            "Администратор портала",
            "admin",
            None,
        ),
        (
            "attestation1@top-academy.ru",
            "Сотрудник аттестации 1",
            "attestation",
            None,
        ),
        (
            "attestation2@top-academy.ru",
            "Сотрудник аттестации 2",
            "attestation",
            None,
        ),
        (
            "attestation3@top-academy.ru",
            "Сотрудник аттестации 3",
            "attestation",
            None,
        ),
        (
            "kazan.director@top-academy.ru",
            "Директор филиала Казань",
            "branch",
            branches["Казань"],
        ),
        (
            "kazan.mup@top-academy.ru",
            "МУП филиала Казань",
            "branch",
            branches["Казань"],
        ),
        (
            "novorossiysk.director@top-academy.ru",
            "Директор филиала Новороссийск",
            "branch",
            branches["Новороссийск"],
        ),
        (
            "novorossiysk.mup@top-academy.ru",
            "МУП филиала Новороссийск",
            "branch",
            branches["Новороссийск"],
        ),
    ]

    for email, full_name, role, branch_id in demo_users:
        db.execute(
            """
            INSERT OR IGNORE INTO users (
                email, password_hash, full_name,
                role, branch_id, active, created_at
            )
            VALUES (?, ?, ?, ?, ?, 1, ?)
            """,
            (
                email,
                generate_password_hash("Pilot123!"),
                full_name,
                role,
                branch_id,
                now,
            ),
        )

    db.commit()
    db.close()


# ============================================================
# Авторизация и безопасность
# ============================================================

@app.before_request
def load_current_user() -> None:
    if "csrf_token" not in session:
        session["csrf_token"] = secrets.token_urlsafe(32)

    user_id = session.get("user_id")
    g.current_user = None

    if user_id:
        g.current_user = get_db().execute(
            """
            SELECT users.*, branches.name AS branch_name
            FROM users
            LEFT JOIN branches ON branches.id = users.branch_id
            WHERE users.id = ? AND users.active = 1
            """,
            (user_id,),
        ).fetchone()

    if request.method == "POST":
        submitted_token = request.form.get("csrf_token", "")
        stored_token = session.get("csrf_token", "")

        if not hmac.compare_digest(submitted_token, stored_token):
            abort(400, "Неверный CSRF-токен.")


def login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not g.current_user:
            return redirect(url_for("login"))
        return view(*args, **kwargs)

    return wrapped


def roles_required(*roles):
    def decorator(view):
        @wraps(view)
        def wrapped(*args, **kwargs):
            if not g.current_user:
                return redirect(url_for("login"))

            if g.current_user["role"] not in roles:
                abort(403)

            return view(*args, **kwargs)

        return wrapped

    return decorator


def audit(action: str, student_id: int | None = None, details: str = "") -> None:
    user_id = g.current_user["id"] if g.current_user else None

    get_db().execute(
        """
        INSERT INTO audit_log (
            user_id, student_id, action, details, created_at
        )
        VALUES (?, ?, ?, ?, ?)
        """,
        (
            user_id,
            student_id,
            action,
            details,
            datetime.now().isoformat(timespec="seconds"),
        ),
    )
    get_db().commit()


def get_student_or_404(student_id: int) -> sqlite3.Row:
    student = get_db().execute(
        """
        SELECT
            students.*,
            branches.name AS branch_name,
            users.full_name AS assigned_name
        FROM students
        JOIN branches ON branches.id = students.branch_id
        LEFT JOIN users ON users.id = students.assigned_to
        WHERE students.id = ?
        """,
        (student_id,),
    ).fetchone()

    if not student:
        abort(404)

    if (
        g.current_user["role"] == "branch"
        and student["branch_id"] != g.current_user["branch_id"]
    ):
        abort(403)

    return student


def full_name(last: str, first: str, middle: str | None) -> str:
    return " ".join(part for part in (last, first, middle) if part)


def calculate_age(birth_date: str) -> int:
    born = date.fromisoformat(birth_date)
    today = date.today()

    return (
        today.year
        - born.year
        - ((today.month, today.day) < (born.month, born.day))
    )


def document_is_required(student: sqlite3.Row, rule: str) -> bool:
    age = calculate_age(student["birth_date"])
    class_number = int(student["class_number"])

    conditions = {
        "all": True,
        "age_14": age >= 14,
        "class_2": class_number >= 2,
        "class_10": class_number >= 10,
        "relation_proof": bool(student["needs_relation_proof"]),
        "russian_citizen": student["citizenship"].strip().upper() == "РФ",
    }

    return conditions.get(rule, False)


def build_document_checklist(student: sqlite3.Row) -> list[dict[str, Any]]:
    latest_documents = {
        row["document_type"]: row
        for row in get_db().execute(
            """
            SELECT d.*
            FROM documents d
            JOIN (
                SELECT document_type, MAX(version) AS latest_version
                FROM documents
                WHERE student_id = ?
                GROUP BY document_type
            ) latest
              ON latest.document_type = d.document_type
             AND latest.latest_version = d.version
            WHERE d.student_id = ?
            """,
            (student["id"], student["id"]),
        ).fetchall()
    }

    checklist = []

    for code, config in DOCUMENT_TYPES.items():
        checklist.append(
            {
                "code": code,
                "name": config["name"],
                "group": config["group"],
                "help": config.get("help", ""),
                "required": document_is_required(student, config["rule"]),
                "document": latest_documents.get(code),
            }
        )

    return checklist


def missing_required_documents(student: sqlite3.Row) -> list[str]:
    return [
        item["name"]
        for item in build_document_checklist(student)
        if item["required"] and not item["document"]
    ]


def allowed_file(filename: str) -> bool:
    return (
        "." in filename
        and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS
    )


# ============================================================
# HTML
# ============================================================

BASE_START = """
<!doctype html>
<html lang="ru">
<head>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <title>{{ title }} — Портал прикрепления</title>
    <style>
        :root {
            --yellow: #ffd500;
            --yellow-light: #fff7c2;
            --black: #171717;
            --gray-50: #f8f8f8;
            --gray-100: #eeeeee;
            --gray-300: #cccccc;
            --gray-600: #666666;
            --white: #ffffff;
            --red: #c62828;
            --red-light: #ffe4e4;
            --green: #147d3f;
            --green-light: #def6e7;
            --blue: #225ea8;
            --blue-light: #e2efff;
        }

        * { box-sizing: border-box; }

        body {
            margin: 0;
            color: var(--black);
            background: #f5f5f5;
            font-family: Arial, Helvetica, sans-serif;
        }

        a { color: inherit; }

        .layout {
            min-height: 100vh;
            display: grid;
            grid-template-columns: 260px 1fr;
        }

        .sidebar {
            background: var(--black);
            color: white;
            padding: 24px 18px;
        }

        .brand {
            display: flex;
            align-items: center;
            gap: 12px;
            margin-bottom: 34px;
        }

        .brand img {
            width: 55px;
            height: 55px;
            object-fit: contain;
            background: white;
            border-radius: 10px;
        }

        .brand-title {
            font-size: 17px;
            line-height: 1.25;
            font-weight: 700;
        }

        .nav-link {
            display: block;
            padding: 12px 14px;
            margin: 4px 0;
            border-radius: 9px;
            color: white;
            text-decoration: none;
        }

        .nav-link:hover {
            color: var(--black);
            background: var(--yellow);
        }

        .user-box {
            margin-top: 30px;
            padding: 14px;
            border: 1px solid #444;
            border-radius: 10px;
            font-size: 13px;
        }

        .main {
            min-width: 0;
        }

        .topbar {
            min-height: 72px;
            display: flex;
            align-items: center;
            justify-content: space-between;
            padding: 16px 30px;
            background: white;
            border-bottom: 1px solid var(--gray-100);
        }

        .content {
            max-width: 1500px;
            padding: 28px 30px 60px;
        }

        h1 { margin: 0 0 22px; }
        h2 { margin-top: 28px; }

        .grid {
            display: grid;
            gap: 18px;
        }

        .metrics {
            grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
        }

        .card {
            padding: 20px;
            background: white;
            border: 1px solid #e7e7e7;
            border-radius: 14px;
            box-shadow: 0 5px 20px rgba(0,0,0,.035);
        }

        .metric-number {
            font-size: 32px;
            font-weight: 800;
            margin-top: 8px;
        }

        .metric-yellow {
            border-top: 5px solid var(--yellow);
        }

        .btn {
            display: inline-block;
            padding: 11px 17px;
            border: 0;
            border-radius: 9px;
            cursor: pointer;
            font-weight: 700;
            text-decoration: none;
        }

        .btn-primary {
            color: var(--black);
            background: var(--yellow);
        }

        .btn-secondary {
            color: var(--black);
            background: white;
            border: 1px solid var(--black);
        }

        .btn-danger {
            color: white;
            background: var(--red);
        }

        .btn-green {
            color: white;
            background: var(--green);
        }

        .btn-small {
            padding: 7px 11px;
            font-size: 13px;
        }

        form.inline {
            display: inline;
        }

        label {
            display: block;
            margin-bottom: 6px;
            font-weight: 700;
        }

        input, select, textarea {
            width: 100%;
            padding: 11px 12px;
            border: 1px solid var(--gray-300);
            border-radius: 8px;
            font: inherit;
            background: white;
        }

        textarea { min-height: 100px; }

        .form-grid {
            display: grid;
            grid-template-columns: repeat(2, minmax(0, 1fr));
            gap: 18px;
        }

        .form-section {
            margin-bottom: 25px;
            padding: 22px;
            background: white;
            border-radius: 14px;
            border: 1px solid #e5e5e5;
        }

        table {
            width: 100%;
            border-collapse: collapse;
            background: white;
        }

        th, td {
            padding: 12px 10px;
            border-bottom: 1px solid var(--gray-100);
            text-align: left;
            vertical-align: top;
        }

        th {
            font-size: 13px;
            background: #fafafa;
        }

        .status {
            display: inline-block;
            padding: 5px 9px;
            border-radius: 20px;
            font-size: 12px;
            font-weight: 700;
        }

        .status.gray { background: #eeeeee; }
        .status.yellow { background: var(--yellow-light); }
        .status.red { color: var(--red); background: var(--red-light); }
        .status.blue { color: var(--blue); background: var(--blue-light); }
        .status.light-green { color: var(--green); background: var(--green-light); }
        .status.green { color: white; background: var(--green); }

        .alert {
            padding: 14px 17px;
            margin-bottom: 18px;
            border-radius: 9px;
        }

        .alert.info { background: var(--yellow-light); }
        .alert.error { color: var(--red); background: var(--red-light); }
        .alert.success { color: var(--green); background: var(--green-light); }

        .document-row {
            display: grid;
            grid-template-columns: minmax(280px, 1fr) 150px 180px 190px;
            align-items: center;
            gap: 14px;
            padding: 15px 0;
            border-bottom: 1px solid var(--gray-100);
        }

        .required { color: var(--red); font-weight: 700; }
        .optional { color: var(--gray-600); }

        .tabs {
            display: flex;
            gap: 8px;
            flex-wrap: wrap;
            margin-bottom: 18px;
        }

        .tab {
            padding: 8px 12px;
            border-radius: 20px;
            background: #eee;
        }

        .login-page {
            min-height: 100vh;
            display: grid;
            place-items: center;
            background:
                linear-gradient(135deg, #ffffff 0%, #fff8c8 100%);
        }

        .login-box {
            width: min(440px, calc(100% - 30px));
            padding: 34px;
            background: white;
            border-radius: 18px;
            border-top: 8px solid var(--yellow);
            box-shadow: 0 20px 50px rgba(0,0,0,.12);
        }

        .login-logo {
            max-width: 160px;
            display: block;
            margin: 0 auto 20px;
        }

        .muted { color: var(--gray-600); }
        .space { margin-top: 18px; }

        @media (max-width: 900px) {
            .layout { grid-template-columns: 1fr; }
            .sidebar { display: none; }
            .form-grid { grid-template-columns: 1fr; }
            .document-row { grid-template-columns: 1fr; }
            .content { padding: 20px 15px 50px; }
        }
    </style>
</head>
<body>
{% if current_user %}
<div class="layout">
    <aside class="sidebar">
        <div class="brand">
            <img src="{{ url_for('logo') }}" alt="Логотип">
            <div class="brand-title">Портал<br>прикрепления</div>
        </div>

        <a class="nav-link" href="{{ url_for('dashboard') }}">Главная</a>
        <a class="nav-link" href="{{ url_for('students') }}">Ученики</a>

        {% if current_user.role in ('attestation', 'admin') %}
            <a class="nav-link" href="{{ url_for('review_queue') }}">
                Очередь проверки
            </a>
            <a class="nav-link" href="{{ url_for('export_page') }}">
                Выгрузка Дневник.ру
            </a>
        {% endif %}

        {% if current_user.role == 'admin' %}
            <a class="nav-link" href="{{ url_for('admin_users') }}">
                Пользователи
            </a>
        {% endif %}

        <div class="user-box">
            <strong>{{ current_user.full_name }}</strong><br>
            {{ current_user.email }}<br>
            {% if current_user.branch_name %}
                {{ current_user.branch_name }}
            {% else %}
                {{ current_user.role }}
            {% endif %}
        </div>
    </aside>

    <main class="main">
        <header class="topbar">
            <strong>{{ title }}</strong>
            <a href="{{ url_for('logout') }}">Выйти</a>
        </header>

        <section class="content">
            {% with messages = get_flashed_messages(with_categories=true) %}
                {% for category, message in messages %}
                    <div class="alert {{ category }}">{{ message }}</div>
                {% endfor %}
            {% endwith %}
"""

BASE_END = """
        </section>
    </main>
</div>
{% else %}
    {{ login_content | safe }}
{% endif %}
</body>
</html>
"""


def render_page(
    title: str,
    body: str,
    *,
    login_content: str = "",
    **context: Any,
):
    return render_template_string(
        BASE_START + body + BASE_END,
        title=title,
        current_user=g.current_user,
        csrf_token=session["csrf_token"],
        status_labels=STATUS_LABELS,
        status_classes=STATUS_CLASSES,
        login_content=login_content,
        **context,
    )


# ============================================================
# Системные страницы
# ============================================================

@app.route("/logo.jpg")
def logo():
    logo_path = BASE_DIR / "logo.jpg"

    if logo_path.exists():
        return send_from_directory(BASE_DIR, "logo.jpg")

    abort(404)


@app.route("/login", methods=["GET", "POST"])
def login():
    if g.current_user:
        return redirect(url_for("dashboard"))

    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")

        user = get_db().execute(
            """
            SELECT *
            FROM users
            WHERE email = ? AND active = 1
            """,
            (email,),
        ).fetchone()

        if (
            not email.endswith("@top-academy.ru")
            or not user
            or not check_password_hash(user["password_hash"], password)
        ):
            flash("Неверная корпоративная почта или пароль.", "error")
        else:
            session.clear()
            session["user_id"] = user["id"]
            session["csrf_token"] = secrets.token_urlsafe(32)

            audit("login", details=email)
            return redirect(url_for("dashboard"))

    login_content = """
    <div class="login-page">
        <div class="login-box">
            <img class="login-logo" src="{{ url_for('logo') }}" alt="Логотип">
            <h1>Портал прикрепления учеников</h1>
            <p class="muted">
                Вход только с корпоративной почтой @top-academy.ru
            </p>

            <form method="post">
                <input type="hidden" name="csrf_token" value="{{ csrf_token }}">

                <div class="space">
                    <label>Корпоративная почта</label>
                    <input
                        type="email"
                        name="email"
                        placeholder="name@top-academy.ru"
                        required
                    >
                </div>

                <div class="space">
                    <label>Пароль</label>
                    <input type="password" name="password" required>
                </div>

                <button class="btn btn-primary space" type="submit">
                    Войти
                </button>
            </form>
        </div>
    </div>
    """

    return render_page(
        "Вход",
        "",
        login_content=render_template_string(
            login_content,
            csrf_token=session["csrf_token"],
        ),
    )


@app.route("/logout")
@login_required
def logout():
    audit("logout")
    session.clear()
    return redirect(url_for("login"))


# ============================================================
# Дашборд
# ============================================================

@app.route("/")
@login_required
def dashboard():
    db = get_db()

    filters = []
    params: list[Any] = []

    if g.current_user["role"] == "branch":
        filters.append("branch_id = ?")
        params.append(g.current_user["branch_id"])

    where_clause = " WHERE " + " AND ".join(filters) if filters else ""

    total = db.execute(
        f"SELECT COUNT(*) FROM students {where_clause}",
        params,
    ).fetchone()[0]

    status_counts = {
        status: db.execute(
            f"""
            SELECT COUNT(*)
            FROM students
            {where_clause}
            {"AND" if filters else "WHERE"} status = ?
            """,
            [*params, status],
        ).fetchone()[0]
        for status in STATUS_LABELS
    }

    # Для общей сводной панели считаем количество
    # уникальных филиалов, по которым есть ученики.
    active_branches_count = 0

    if g.current_user["role"] != "branch":
        active_branches_count = db.execute(
            """
            SELECT COUNT(DISTINCT branch_id)
            FROM students
            """
        ).fetchone()[0]

    recent = db.execute(
        f"""
        SELECT
            students.*,
            branches.name AS branch_name
        FROM students
        JOIN branches ON branches.id = students.branch_id
        {where_clause}
        ORDER BY students.updated_at DESC
        LIMIT 10
        """,
        params,
    ).fetchall()

    body = """
    <div style="display:flex;justify-content:space-between;gap:15px;align-items:center">
        <div>
            <h1>
                {% if current_user.branch_name %}
                    Филиал {{ current_user.branch_name }}
                {% else %}
                    Сводная панель
                {% endif %}
            </h1>
            <p class="muted">Контроль подготовки комплектов учеников</p>
        </div>

        {% if current_user.role == 'branch' %}
            <a class="btn btn-primary" href="{{ url_for('new_student') }}">
                + Добавить ученика
            </a>
        {% endif %}
    </div>

    <div class="grid metrics">
        <div class="card metric-yellow">
            Всего учеников
            <div class="metric-number">{{ total }}</div>
        </div>

        {% if current_user.role != 'branch' %}
            <div class="card metric-yellow">
                Филиалов с учениками
                <div class="metric-number">
                    {{ active_branches_count }}
                </div>

                <div class="muted" style="margin-top:6px">
                    Уникальные филиалы, по которым
                    заведены карточки учеников
                </div>
            </div>
        {% endif %}

        {% for code, label in status_labels.items() %}
            <div class="card">
                {{ label }}
                <div class="metric-number">{{ status_counts[code] }}</div>
            </div>
        {% endfor %}
    </div>

    <h2>Последние изменения</h2>

    <div class="card">
        <table>
            <thead>
                <tr>
                    <th>Ученик</th>
                    {% if current_user.role != 'branch' %}
                        <th>Филиал</th>
                    {% endif %}
                    <th>Класс</th>
                    <th>Статус</th>
                    <th>Обновлено</th>
                </tr>
            </thead>
            <tbody>
                {% for student in recent %}
                    <tr>
                        <td>
                            <a href="{{ url_for('student_detail', student_id=student.id) }}">
                                {{ student.last_name }}
                                {{ student.first_name }}
                                {{ student.middle_name or '' }}
                            </a>
                        </td>

                        {% if current_user.role != 'branch' %}
                            <td>{{ student.branch_name }}</td>
                        {% endif %}

                        <td>{{ student.class_number }}</td>

                        <td>
                            <span class="status {{ status_classes[student.status] }}">
                                {{ status_labels[student.status] }}
                            </span>
                        </td>

                        <td>{{ student.updated_at[:16].replace('T', ' ') }}</td>
                    </tr>
                {% else %}
                    <tr><td colspan="5">Карточек пока нет.</td></tr>
                {% endfor %}
            </tbody>
        </table>
    </div>
    """

    return render_page(
        "Главная",
        body,
        total=total,
        status_counts=status_counts,
        active_branches_count=active_branches_count,
        recent=recent,
    )


# ============================================================
# Ученики
# ============================================================

@app.route("/students")
@login_required
def students():
    db = get_db()

    query = """
        SELECT
            students.*,
            branches.name AS branch_name,
            (
                SELECT COUNT(*)
                FROM comments
                WHERE comments.student_id = students.id
                  AND comments.is_open = 1
            ) AS open_comments
        FROM students
        JOIN branches ON branches.id = students.branch_id
        WHERE 1 = 1
    """

    params: list[Any] = []

    if g.current_user["role"] == "branch":
        query += " AND students.branch_id = ?"
        params.append(g.current_user["branch_id"])

    search = request.args.get("search", "").strip()
    status = request.args.get("status", "").strip()
    class_number = request.args.get("class_number", "").strip()
    branch_id = request.args.get("branch_id", "").strip()

    if search:
        query += """
            AND (
                students.last_name LIKE ?
                OR students.first_name LIKE ?
                OR students.student_snils LIKE ?
            )
        """
        pattern = f"%{search}%"
        params.extend([pattern, pattern, pattern])

    if status in STATUS_LABELS:
        query += " AND students.status = ?"
        params.append(status)

    if (
        class_number.isdigit()
        and 1 <= int(class_number) <= 11
    ):
        query += " AND students.class_number = ?"
        params.append(int(class_number))

    # Администратор может отфильтровать учеников
    # по конкретному филиалу.
    if (
        g.current_user["role"] != "branch"
        and branch_id.isdigit()
    ):
        branch_exists = db.execute(
            """
            SELECT id
            FROM branches
            WHERE id = ?
            """,
            (int(branch_id),),
        ).fetchone()

        if branch_exists:
            query += " AND students.branch_id = ?"
            params.append(int(branch_id))

    query += " ORDER BY students.updated_at DESC"

    rows = db.execute(query, params).fetchall()

    # Список филиалов нужен для фильтра администратора.
    branches = []

    if g.current_user["role"] != "branch":
        branches = db.execute(
            """
            SELECT id, name
            FROM branches
            ORDER BY name
            """
        ).fetchall()

    body = """
    <div style="display:flex;justify-content:space-between;align-items:center;gap:20px">
        <h1>Ученики</h1>

        {% if current_user.role == 'branch' %}
            <a class="btn btn-primary" href="{{ url_for('new_student') }}">
                + Добавить ученика
            </a>
        {% endif %}
    </div>

    <form class="card form-grid" method="get">
        <div>
            <label>Поиск</label>
            <input
                name="search"
                value="{{ request.args.get('search', '') }}"
                placeholder="ФИО или СНИЛС"
            >
        </div>

        <div>
            <label>Статус</label>
            <select name="status">
                <option value="">Все статусы</option>
                {% for code, label in status_labels.items() %}
                    <option
                        value="{{ code }}"
                        {% if request.args.get('status') == code %}selected{% endif %}
                    >
                        {{ label }}
                    </option>
                {% endfor %}
            </select>
        </div>

        {% if current_user.role != 'branch' %}
            <div>
                <label>Филиал</label>

                <select name="branch_id">
                    <option value="">
                        Все филиалы
                    </option>

                    {% for branch in branches %}
                        <option
                            value="{{ branch.id }}"
                            {% if
                                request.args.get(
                                    'branch_id',
                                    ''
                                )
                                == branch.id|string
                            %}
                                selected
                            {% endif %}
                        >
                            {{ branch.name }}
                        </option>
                    {% endfor %}
                </select>
            </div>
        {% endif %}

        <div>
            <label>Класс</label>

            <select name="class_number">
                <option value="">
                    Все классы
                </option>

                {% for class_value in range(1, 12) %}
                    <option
                        value="{{ class_value }}"
                        {% if
                            request.args.get(
                                'class_number',
                                ''
                            )
                            == class_value|string
                        %}
                            selected
                        {% endif %}
                    >
                        {{ class_value }} класс
                    </option>
                {% endfor %}
            </select>
        </div>

        <div
            style="
                align-self:end;
                display:flex;
                gap:10px;
                flex-wrap:wrap;
            "
        >
            <button class="btn btn-primary">
                Применить
            </button>

            <a
                class="btn btn-secondary"
                href="{{ url_for('students') }}"
            >
                Сбросить
            </a>
        </div>
    </form>

    <div class="card space">
        <table>
            <thead>
                <tr>
                    <th>Ученик</th>
                    {% if current_user.role != 'branch' %}
                        <th>Филиал</th>
                    {% endif %}
                    <th>Класс</th>
                    <th>Возраст</th>
                    <th>Статус</th>
                    <th>Замечания</th>
                </tr>
            </thead>
            <tbody>
                {% for student in rows %}
                    <tr>
                        <td>
                            <a href="{{ url_for('student_detail', student_id=student.id) }}">
                                {{ student.last_name }}
                                {{ student.first_name }}
                                {{ student.middle_name or '' }}
                            </a>
                        </td>

                        {% if current_user.role != 'branch' %}
                            <td>{{ student.branch_name }}</td>
                        {% endif %}

                        <td>{{ student.class_number }}</td>
                        <td>{{ calculate_age(student.birth_date) }}</td>

                        <td>
                            <span class="status {{ status_classes[student.status] }}">
                                {{ status_labels[student.status] }}
                            </span>
                        </td>

                        <td>{{ student.open_comments }}</td>
                    </tr>
                {% else %}
                    <tr><td colspan="6">Ученики не найдены.</td></tr>
                {% endfor %}
            </tbody>
        </table>
    </div>
    """

    return render_page(
        "Ученики",
        body,
        rows=rows,
        branches=branches,
        calculate_age=calculate_age,
        range=range,
    )


@app.route("/students/new", methods=["GET", "POST"])
@roles_required("branch")
def new_student():
    if request.method == "POST":
        form = request.form
        now = datetime.now().isoformat(timespec="seconds")

        required_fields = [
            "last_name",
            "first_name",
            "birth_date",
            "class_number",
            "parent_last_name",
            "parent_first_name",
            "relation_type",
        ]

        if any(not form.get(field, "").strip() for field in required_fields):
            flash("Заполнены не все обязательные поля.", "error")
        else:
            duplicate = get_db().execute(
                """
                SELECT id
                FROM students
                WHERE branch_id = ?
                  AND lower(last_name) = lower(?)
                  AND lower(first_name) = lower(?)
                  AND birth_date = ?
                """,
                (
                    g.current_user["branch_id"],
                    form["last_name"].strip(),
                    form["first_name"].strip(),
                    form["birth_date"],
                ),
            ).fetchone()

            if duplicate:
                flash(
                    "В филиале уже есть ученик с таким ФИО и датой рождения.",
                    "error",
                )
            else:
                cursor = get_db().execute(
                    """
                    INSERT INTO students (
                        branch_id,
                        created_by,

                        last_name,
                        first_name,
                        middle_name,
                        gender,
                        birth_date,
                        citizenship,
                        class_number,

                        arrival_school_date,
                        arrival_class_date,

                        student_snils,
                        birth_certificate_series,
                        birth_certificate_number,
                        child_passport_series,
                        child_passport_number,
                        student_phone,
                        student_email,

                        parent_last_name,
                        parent_first_name,
                        parent_middle_name,
                        parent_gender,
                        parent_birth_date,
                        parent_snils,
                        parent_passport_series,
                        parent_passport_number,
                        relation_type,
                        parent_phone,
                        parent_email,

                        needs_relation_proof,
                        status,
                        created_at,
                        updated_at
                    )
                    VALUES (
                        ?, ?,
                        ?, ?, ?, ?, ?, ?, ?,
                        ?, ?,
                        ?, ?, ?, ?, ?, ?, ?,
                        ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?,
                        ?, 'draft', ?, ?
                    )
                    """,
                    (
                        g.current_user["branch_id"],
                        g.current_user["id"],

                        form["last_name"].strip(),
                        form["first_name"].strip(),
                        form.get("middle_name", "").strip(),
                        form.get("gender", ""),
                        form["birth_date"],
                        form.get("citizenship", "РФ").strip(),
                        int(form["class_number"]),

                        form.get("arrival_school_date") or None,
                        form.get("arrival_class_date") or None,

                        form.get("student_snils", "").strip(),
                        form.get("birth_certificate_series", "").strip(),
                        form.get("birth_certificate_number", "").strip(),
                        form.get("child_passport_series", "").strip(),
                        form.get("child_passport_number", "").strip(),
                        form.get("student_phone", "").strip(),
                        form.get("student_email", "").strip(),

                        form["parent_last_name"].strip(),
                        form["parent_first_name"].strip(),
                        form.get("parent_middle_name", "").strip(),
                        form.get("parent_gender", ""),
                        form.get("parent_birth_date") or None,
                        form.get("parent_snils", "").strip(),
                        form.get("parent_passport_series", "").strip(),
                        form.get("parent_passport_number", "").strip(),
                        form["relation_type"].strip(),
                        form.get("parent_phone", "").strip(),
                        form.get("parent_email", "").strip(),

                        1 if form.get("needs_relation_proof") else 0,
                        now,
                        now,
                    ),
                )

                student_id = cursor.lastrowid
                get_db().commit()

                audit(
                    "student_created",
                    student_id,
                    "Создана карточка ученика",
                )

                flash("Карточка ученика создана.", "success")
                return redirect(
                    url_for("student_detail", student_id=student_id)
                )

    body = """
    <h1>Новый ученик</h1>

    <form method="post">
        <input type="hidden" name="csrf_token" value="{{ csrf_token }}">

        <section class="form-section">
            <h2>Сведения об ученике</h2>

            <div class="form-grid">
                <div>
                    <label>Фамилия *</label>
                    <input name="last_name" required>
                </div>

                <div>
                    <label>Имя *</label>
                    <input name="first_name" required>
                </div>

                <div>
                    <label>Отчество</label>
                    <input name="middle_name">
                </div>

                <div>
                    <label>Пол</label>
                    <select name="gender">
                        <option value="">Не выбран</option>
                        <option>Мужской</option>
                        <option>Женский</option>
                    </select>
                </div>

                <div>
                    <label>Дата рождения *</label>
                    <input type="date" name="birth_date" required>
                </div>

                <div>
                    <label>Гражданство *</label>
                    <input name="citizenship" value="РФ" required>
                </div>

                <div>
                    <label>Класс поступления *</label>
                    <input
                        type="number"
                        name="class_number"
                        min="1"
                        max="11"
                        required
                    >
                </div>

                <div>
                    <label>СНИЛС ученика</label>
                    <input name="student_snils" placeholder="000-000-000 00">
                </div>

                <div>
                    <label>Дата прибытия в ОО</label>
                    <input type="date" name="arrival_school_date">
                </div>

                <div>
                    <label>Дата прибытия в класс</label>
                    <input type="date" name="arrival_class_date">
                </div>

                <div>
                    <label>Серия свидетельства о рождении</label>
                    <input name="birth_certificate_series">
                </div>

                <div>
                    <label>Номер свидетельства о рождении</label>
                    <input name="birth_certificate_number">
                </div>

                <div>
                    <label>Серия паспорта ученика</label>
                    <input name="child_passport_series">
                </div>

                <div>
                    <label>Номер паспорта ученика</label>
                    <input name="child_passport_number">
                </div>

                <div>
                    <label>Телефон ученика</label>
                    <input name="student_phone">
                </div>

                <div>
                    <label>Email ученика</label>
                    <input type="email" name="student_email">
                </div>
            </div>
        </section>

        <section class="form-section">
            <h2>Законный представитель</h2>

            <div class="alert info">
                Во всех заявлениях, согласиях и документах должен быть
                указан один и тот же законный представитель.
            </div>

            <div class="form-grid">
                <div>
                    <label>Фамилия *</label>
                    <input name="parent_last_name" required>
                </div>

                <div>
                    <label>Имя *</label>
                    <input name="parent_first_name" required>
                </div>

                <div>
                    <label>Отчество</label>
                    <input name="parent_middle_name">
                </div>

                <div>
                    <label>Пол</label>
                    <select name="parent_gender">
                        <option value="">Не выбран</option>
                        <option>Мужской</option>
                        <option>Женский</option>
                    </select>
                </div>

                <div>
                    <label>Дата рождения</label>
                    <input type="date" name="parent_birth_date">
                </div>

                <div>
                    <label>Вид родственной связи *</label>
                    <select name="relation_type" required>
                        <option value="">Выберите</option>
                        <option>Мать</option>
                        <option>Отец</option>
                        <option>Опекун</option>
                        <option>Усыновитель</option>
                        <option>Иной законный представитель</option>
                    </select>
                </div>

                <div>
                    <label>СНИЛС</label>
                    <input name="parent_snils">
                </div>

                <div>
                    <label>Серия паспорта</label>
                    <input name="parent_passport_series">
                </div>

                <div>
                    <label>Номер паспорта</label>
                    <input name="parent_passport_number">
                </div>

                <div>
                    <label>Телефон</label>
                    <input name="parent_phone">
                </div>

                <div>
                    <label>Email</label>
                    <input type="email" name="parent_email">
                </div>
            </div>

            <p>
                <label>
                    <input
                        style="width:auto"
                        type="checkbox"
                        name="needs_relation_proof"
                    >
                    Нужен документ о смене фамилии, опеке или усыновлении
                </label>
            </p>
        </section>

        <button class="btn btn-primary" type="submit">
            Создать карточку
        </button>
    </form>
    """

    return render_page("Новый ученик", body)



# ============================================================
# NAME EDIT V9.2.1
# Редактирование ФИО после создания карточки
# ============================================================


# ============================================================
# FULL CARD EDIT V9.3
# Редактирование всех данных карточки
# ============================================================

@app.route(
    "/students/<int:student_id>/edit-names",
    methods=["GET", "POST"],
)
@login_required
def edit_student_names_v9_2_1(student_id: int):

    student = get_student_or_404(
        student_id
    )

    role = g.current_user["role"]

    # Филиал может редактировать
    # только собственных учеников.
    if (
        role == "branch"
        and student["branch_id"]
        != g.current_user["branch_id"]
    ):
        abort(403)

    if role not in (
        "branch",
        "attestation",
        "admin",
    ):
        abort(403)

    if request.method == "POST":

        form = request.form

        required_fields = {
            "last_name":
                "Фамилия ученика",

            "first_name":
                "Имя ученика",

            "birth_date":
                "Дата рождения ученика",

            "citizenship":
                "Гражданство",

            "class_number":
                "Класс поступления",

            "parent_last_name":
                "Фамилия законного представителя",

            "parent_first_name":
                "Имя законного представителя",

            "relation_type":
                "Вид родственной связи",
        }

        errors = []

        for field, label in (
            required_fields.items()
        ):
            if not form.get(
                field,
                "",
            ).strip():

                errors.append(
                    f"Заполните поле "
                    f"«{label}»."
                )

        # --------------------------------
        # Класс
        # --------------------------------

        class_number = None

        try:
            class_number = int(
                form.get(
                    "class_number",
                    "",
                )
            )

            if not (
                1 <= class_number <= 11
            ):
                raise ValueError

        except ValueError:
            errors.append(
                "Класс должен быть "
                "от 1 до 11."
            )

        # --------------------------------
        # Даты
        # --------------------------------

        date_fields = {
            "birth_date":
                "Дата рождения ученика",

            "arrival_school_date":
                "Дата прибытия в ОО",

            "arrival_class_date":
                "Дата прибытия в класс",

            "parent_birth_date":
                "Дата рождения "
                "законного представителя",
        }

        for field, label in (
            date_fields.items()
        ):

            value = form.get(
                field,
                "",
            ).strip()

            if not value:
                continue

            try:
                datetime.strptime(
                    value,
                    "%Y-%m-%d",
                )

            except ValueError:
                errors.append(
                    f"Некорректная дата: "
                    f"«{label}»."
                )

        # --------------------------------
        # Справочники
        # --------------------------------

        gender = form.get(
            "gender",
            "",
        ).strip()

        parent_gender = form.get(
            "parent_gender",
            "",
        ).strip()

        if gender not in (
            "",
            "Мужской",
            "Женский",
        ):
            errors.append(
                "Некорректно указан "
                "пол ученика."
            )

        if parent_gender not in (
            "",
            "Мужской",
            "Женский",
        ):
            errors.append(
                "Некорректно указан "
                "пол законного представителя."
            )

        relation_type = form.get(
            "relation_type",
            "",
        ).strip()

        allowed_relations = (
            "Мать",
            "Отец",
            "Опекун",
            "Усыновитель",
            "Иной законный представитель",
        )

        if (
            relation_type
            and relation_type
            not in allowed_relations
        ):
            errors.append(
                "Некорректно указан "
                "вид родственной связи."
            )

        # --------------------------------
        # Проверяем дубль после изменения
        # ФИО / даты рождения.
        # --------------------------------

        if not errors:

            duplicate = (
                get_db().execute(
                    """
                    SELECT id
                    FROM students
                    WHERE branch_id = ?
                      AND id <> ?
                      AND lower(last_name)
                            = lower(?)
                      AND lower(first_name)
                            = lower(?)
                      AND birth_date = ?
                    LIMIT 1
                    """,
                    (
                        student[
                            "branch_id"
                        ],
                        student_id,
                        form[
                            "last_name"
                        ].strip(),
                        form[
                            "first_name"
                        ].strip(),
                        form[
                            "birth_date"
                        ].strip(),
                    ),
                ).fetchone()
            )

            if duplicate:
                errors.append(
                    "В филиале уже есть "
                    "другой ученик с таким "
                    "ФИО и датой рождения."
                )

        if errors:

            for error in errors:
                flash(
                    error,
                    "error",
                )

            return redirect(
                request.url
            )

        # ====================================================
        # Все редактируемые поля
        # ====================================================

        values = {

            # Ученик
            "last_name":
                form[
                    "last_name"
                ].strip(),

            "first_name":
                form[
                    "first_name"
                ].strip(),

            "middle_name":
                form.get(
                    "middle_name",
                    "",
                ).strip(),

            "gender":
                gender,

            "birth_date":
                form[
                    "birth_date"
                ].strip(),

            "citizenship":
                form[
                    "citizenship"
                ].strip(),

            "class_number":
                class_number,

            "student_snils":
                form.get(
                    "student_snils",
                    "",
                ).strip(),

            "arrival_school_date":
                (
                    form.get(
                        "arrival_school_date",
                        "",
                    ).strip()
                    or None
                ),

            "arrival_class_date":
                (
                    form.get(
                        "arrival_class_date",
                        "",
                    ).strip()
                    or None
                ),

            "birth_certificate_series":
                form.get(
                    "birth_certificate_series",
                    "",
                ).strip(),

            "birth_certificate_number":
                form.get(
                    "birth_certificate_number",
                    "",
                ).strip(),

            "child_passport_series":
                form.get(
                    "child_passport_series",
                    "",
                ).strip(),

            "child_passport_number":
                form.get(
                    "child_passport_number",
                    "",
                ).strip(),

            "student_phone":
                form.get(
                    "student_phone",
                    "",
                ).strip(),

            "student_email":
                form.get(
                    "student_email",
                    "",
                ).strip(),

            # Законный представитель
            "parent_last_name":
                form[
                    "parent_last_name"
                ].strip(),

            "parent_first_name":
                form[
                    "parent_first_name"
                ].strip(),

            "parent_middle_name":
                form.get(
                    "parent_middle_name",
                    "",
                ).strip(),

            "parent_gender":
                parent_gender,

            "parent_birth_date":
                (
                    form.get(
                        "parent_birth_date",
                        "",
                    ).strip()
                    or None
                ),

            "parent_snils":
                form.get(
                    "parent_snils",
                    "",
                ).strip(),

            "parent_passport_series":
                form.get(
                    "parent_passport_series",
                    "",
                ).strip(),

            "parent_passport_number":
                form.get(
                    "parent_passport_number",
                    "",
                ).strip(),

            "relation_type":
                relation_type,

            "parent_phone":
                form.get(
                    "parent_phone",
                    "",
                ).strip(),

            "parent_email":
                form.get(
                    "parent_email",
                    "",
                ).strip(),

            "needs_relation_proof":
                (
                    1
                    if form.get(
                        "needs_relation_proof"
                    )
                    else 0
                ),
        }

        # --------------------------------
        # Формируем аудит только по
        # реально измененным значениям.
        # --------------------------------

        field_labels = {

            "last_name":
                "Фамилия ученика",

            "first_name":
                "Имя ученика",

            "middle_name":
                "Отчество ученика",

            "gender":
                "Пол ученика",

            "birth_date":
                "Дата рождения ученика",

            "citizenship":
                "Гражданство",

            "class_number":
                "Класс",

            "student_snils":
                "СНИЛС ученика",

            "arrival_school_date":
                "Дата прибытия в ОО",

            "arrival_class_date":
                "Дата прибытия в класс",

            "birth_certificate_series":
                "Серия свидетельства",

            "birth_certificate_number":
                "Номер свидетельства",

            "child_passport_series":
                "Серия паспорта ученика",

            "child_passport_number":
                "Номер паспорта ученика",

            "student_phone":
                "Телефон ученика",

            "student_email":
                "Email ученика",

            "parent_last_name":
                "Фамилия представителя",

            "parent_first_name":
                "Имя представителя",

            "parent_middle_name":
                "Отчество представителя",

            "parent_gender":
                "Пол представителя",

            "parent_birth_date":
                "Дата рождения представителя",

            "parent_snils":
                "СНИЛС представителя",

            "parent_passport_series":
                "Серия паспорта представителя",

            "parent_passport_number":
                "Номер паспорта представителя",

            "relation_type":
                "Родственная связь",

            "parent_phone":
                "Телефон представителя",

            "parent_email":
                "Email представителя",

            "needs_relation_proof":
                "Нужен подтверждающий документ",
        }

        changes = []

        for field, new_value in (
            values.items()
        ):

            old_value = student[field]

            old_normalized = (
                ""
                if old_value is None
                else str(old_value)
            )

            new_normalized = (
                ""
                if new_value is None
                else str(new_value)
            )

            if (
                old_normalized
                != new_normalized
            ):
                changes.append(
                    (
                        f"{field_labels[field]}: "
                        f"«{old_normalized}» → "
                        f"«{new_normalized}»"
                    )
                )

        if not changes:

            flash(
                "Изменений нет.",
                "info",
            )

            return redirect(
                url_for(
                    "student_detail",
                    student_id=student_id,
                )
            )

        timestamp = (
            datetime.now()
            .isoformat(
                timespec="seconds"
            )
        )

        # ====================================================
        # Сохраняем карточку
        # ====================================================

        get_db().execute(
            """
            UPDATE students
            SET
                last_name = ?,
                first_name = ?,
                middle_name = ?,
                gender = ?,
                birth_date = ?,
                citizenship = ?,
                class_number = ?,

                student_snils = ?,
                arrival_school_date = ?,
                arrival_class_date = ?,

                birth_certificate_series = ?,
                birth_certificate_number = ?,

                child_passport_series = ?,
                child_passport_number = ?,

                student_phone = ?,
                student_email = ?,

                parent_last_name = ?,
                parent_first_name = ?,
                parent_middle_name = ?,
                parent_gender = ?,
                parent_birth_date = ?,

                parent_snils = ?,
                parent_passport_series = ?,
                parent_passport_number = ?,

                relation_type = ?,
                parent_phone = ?,
                parent_email = ?,

                needs_relation_proof = ?,

                updated_at = ?

            WHERE id = ?
            """,
            (
                values[
                    "last_name"
                ],
                values[
                    "first_name"
                ],
                values[
                    "middle_name"
                ],
                values[
                    "gender"
                ],
                values[
                    "birth_date"
                ],
                values[
                    "citizenship"
                ],
                values[
                    "class_number"
                ],

                values[
                    "student_snils"
                ],
                values[
                    "arrival_school_date"
                ],
                values[
                    "arrival_class_date"
                ],

                values[
                    "birth_certificate_series"
                ],
                values[
                    "birth_certificate_number"
                ],

                values[
                    "child_passport_series"
                ],
                values[
                    "child_passport_number"
                ],

                values[
                    "student_phone"
                ],
                values[
                    "student_email"
                ],

                values[
                    "parent_last_name"
                ],
                values[
                    "parent_first_name"
                ],
                values[
                    "parent_middle_name"
                ],
                values[
                    "parent_gender"
                ],
                values[
                    "parent_birth_date"
                ],

                values[
                    "parent_snils"
                ],
                values[
                    "parent_passport_series"
                ],
                values[
                    "parent_passport_number"
                ],

                values[
                    "relation_type"
                ],
                values[
                    "parent_phone"
                ],
                values[
                    "parent_email"
                ],

                values[
                    "needs_relation_proof"
                ],

                timestamp,
                student_id,
            ),
        )

        # --------------------------------
        # Черновик справки должен брать
        # актуальные данные карточки.
        # Уже выданные документы не меняем.
        # --------------------------------

        try:
            get_db().execute(
                """
                UPDATE
                    post_enrollment_documents

                SET
                    name_override = NULL,
                    updated_at = ?,
                    updated_by = ?

                WHERE
                    student_id = ?
                    AND status = 'draft'
                """,
                (
                    timestamp,
                    g.current_user[
                        "id"
                    ],
                    student_id,
                ),
            )

        except Exception:
            pass

        get_db().commit()

        audit(
            "student_card_updated",
            student_id,
            "; ".join(
                changes
            ),
        )

        flash(
            (
                "Данные карточки "
                "успешно обновлены."
            ),
            "success",
        )

        return redirect(
            url_for(
                "student_detail",
                student_id=student_id,
            )
        )

    # ========================================================
    # Форма
    # ========================================================

    body = """
    <div
        style="
            display:flex;
            justify-content:space-between;
            align-items:center;
            gap:15px;
            flex-wrap:wrap;
        "
    >
        <div>
            <h1 style="margin-bottom:6px">
                Редактирование данных
            </h1>

            <div class="muted">
                {{
                    student.last_name
                }}
                {{
                    student.first_name
                }}
                {{
                    student.middle_name
                    or ''
                }}
            </div>
        </div>

        <a
            class="btn btn-secondary"
            href="{{ url_for(
                'student_detail',
                student_id=student.id
            ) }}"
        >
            Назад к карточке
        </a>
    </div>

    <div class="alert info space">
        После сохранения новые данные
        сразу будут использоваться
        при проверке документов,
        расчете обязательного комплекта
        и формировании Excel-выгрузки.
    </div>

    <form method="post">

        <input
            type="hidden"
            name="csrf_token"
            value="{{ csrf_token }}"
        >

        <!-- ============================================= -->
        <!-- УЧЕНИК -->
        <!-- ============================================= -->

        <section class="form-section">

            <h2 style="margin-top:0">
                Сведения об ученике
            </h2>

            <div class="form-grid">

                <div>
                    <label>
                        Фамилия *
                    </label>

                    <input
                        name="last_name"
                        value="{{
                            student.last_name
                        }}"
                        required
                    >
                </div>

                <div>
                    <label>
                        Имя *
                    </label>

                    <input
                        name="first_name"
                        value="{{
                            student.first_name
                        }}"
                        required
                    >
                </div>

                <div>
                    <label>
                        Отчество
                    </label>

                    <input
                        name="middle_name"
                        value="{{
                            student.middle_name
                            or ''
                        }}"
                    >
                </div>

                <div>
                    <label>
                        Пол
                    </label>

                    <select name="gender">

                        <option
                            value=""
                            {% if
                                not student.gender
                            %}
                                selected
                            {% endif %}
                        >
                            Не выбран
                        </option>

                        <option
                            value="Мужской"
                            {% if
                                student.gender
                                == 'Мужской'
                            %}
                                selected
                            {% endif %}
                        >
                            Мужской
                        </option>

                        <option
                            value="Женский"
                            {% if
                                student.gender
                                == 'Женский'
                            %}
                                selected
                            {% endif %}
                        >
                            Женский
                        </option>

                    </select>
                </div>

                <div>
                    <label>
                        Дата рождения *
                    </label>

                    <input
                        type="date"
                        name="birth_date"
                        value="{{
                            student.birth_date
                        }}"
                        required
                    >
                </div>

                <div>
                    <label>
                        Гражданство *
                    </label>

                    <input
                        name="citizenship"
                        value="{{
                            student.citizenship
                        }}"
                        required
                    >
                </div>

                <div>
                    <label>
                        Класс поступления *
                    </label>

                    <input
                        type="number"
                        name="class_number"
                        min="1"
                        max="11"
                        value="{{
                            student.class_number
                        }}"
                        required
                    >
                </div>

                <div>
                    <label>
                        СНИЛС ученика
                    </label>

                    <input
                        name="student_snils"
                        value="{{
                            student.student_snils
                            or ''
                        }}"
                        placeholder="
                            000-000-000 00
                        "
                    >
                </div>

                <div>
                    <label>
                        Дата прибытия в ОО
                    </label>

                    <input
                        type="date"
                        name="arrival_school_date"
                        value="{{
                            student.arrival_school_date
                            or ''
                        }}"
                    >
                </div>

                <div>
                    <label>
                        Дата прибытия в класс
                    </label>

                    <input
                        type="date"
                        name="arrival_class_date"
                        value="{{
                            student.arrival_class_date
                            or ''
                        }}"
                    >
                </div>

                <div>
                    <label>
                        Серия свидетельства
                        о рождении
                    </label>

                    <input
                        name="
                            birth_certificate_series
                        "
                        value="{{
                            student
                            .birth_certificate_series
                            or ''
                        }}"
                    >
                </div>

                <div>
                    <label>
                        Номер свидетельства
                        о рождении
                    </label>

                    <input
                        name="
                            birth_certificate_number
                        "
                        value="{{
                            student
                            .birth_certificate_number
                            or ''
                        }}"
                    >
                </div>

                <div>
                    <label>
                        Серия паспорта ученика
                    </label>

                    <input
                        name="
                            child_passport_series
                        "
                        value="{{
                            student
                            .child_passport_series
                            or ''
                        }}"
                    >
                </div>

                <div>
                    <label>
                        Номер паспорта ученика
                    </label>

                    <input
                        name="
                            child_passport_number
                        "
                        value="{{
                            student
                            .child_passport_number
                            or ''
                        }}"
                    >
                </div>

                <div>
                    <label>
                        Телефон ученика
                    </label>

                    <input
                        name="student_phone"
                        value="{{
                            student.student_phone
                            or ''
                        }}"
                    >
                </div>

                <div>
                    <label>
                        Email ученика
                    </label>

                    <input
                        type="email"
                        name="student_email"
                        value="{{
                            student.student_email
                            or ''
                        }}"
                    >
                </div>

            </div>

        </section>

        <!-- ============================================= -->
        <!-- ПРЕДСТАВИТЕЛЬ -->
        <!-- ============================================= -->

        <section class="form-section">

            <h2 style="margin-top:0">
                Законный представитель
            </h2>

            <div class="alert info">
                Во всех заявлениях,
                согласиях и документах
                должен быть указан один
                и тот же законный
                представитель.
            </div>

            <div class="form-grid">

                <div>
                    <label>
                        Фамилия *
                    </label>

                    <input
                        name="parent_last_name"
                        value="{{
                            student.parent_last_name
                        }}"
                        required
                    >
                </div>

                <div>
                    <label>
                        Имя *
                    </label>

                    <input
                        name="parent_first_name"
                        value="{{
                            student.parent_first_name
                        }}"
                        required
                    >
                </div>

                <div>
                    <label>
                        Отчество
                    </label>

                    <input
                        name="parent_middle_name"
                        value="{{
                            student
                            .parent_middle_name
                            or ''
                        }}"
                    >
                </div>

                <div>
                    <label>
                        Пол
                    </label>

                    <select
                        name="parent_gender"
                    >

                        <option
                            value=""
                            {% if
                                not student
                                .parent_gender
                            %}
                                selected
                            {% endif %}
                        >
                            Не выбран
                        </option>

                        <option
                            value="Мужской"
                            {% if
                                student
                                .parent_gender
                                == 'Мужской'
                            %}
                                selected
                            {% endif %}
                        >
                            Мужской
                        </option>

                        <option
                            value="Женский"
                            {% if
                                student
                                .parent_gender
                                == 'Женский'
                            %}
                                selected
                            {% endif %}
                        >
                            Женский
                        </option>

                    </select>
                </div>

                <div>
                    <label>
                        Дата рождения
                    </label>

                    <input
                        type="date"
                        name="parent_birth_date"
                        value="{{
                            student
                            .parent_birth_date
                            or ''
                        }}"
                    >
                </div>

                <div>
                    <label>
                        Вид родственной связи *
                    </label>

                    <select
                        name="relation_type"
                        required
                    >

                        {% for relation in [
                            'Мать',
                            'Отец',
                            'Опекун',
                            'Усыновитель',
                            'Иной законный представитель'
                        ] %}

                            <option
                                value="{{ relation }}"
                                {% if
                                    student
                                    .relation_type
                                    == relation
                                %}
                                    selected
                                {% endif %}
                            >
                                {{ relation }}
                            </option>

                        {% endfor %}

                    </select>
                </div>

                <div>
                    <label>
                        СНИЛС
                    </label>

                    <input
                        name="parent_snils"
                        value="{{
                            student.parent_snils
                            or ''
                        }}"
                        placeholder="
                            000-000-000 00
                        "
                    >
                </div>

                <div>
                    <label>
                        Серия паспорта
                    </label>

                    <input
                        name="
                            parent_passport_series
                        "
                        value="{{
                            student
                            .parent_passport_series
                            or ''
                        }}"
                    >
                </div>

                <div>
                    <label>
                        Номер паспорта
                    </label>

                    <input
                        name="
                            parent_passport_number
                        "
                        value="{{
                            student
                            .parent_passport_number
                            or ''
                        }}"
                    >
                </div>

                <div>
                    <label>
                        Телефон
                    </label>

                    <input
                        name="parent_phone"
                        value="{{
                            student.parent_phone
                            or ''
                        }}"
                    >
                </div>

                <div>
                    <label>
                        Email
                    </label>

                    <input
                        type="email"
                        name="parent_email"
                        value="{{
                            student.parent_email
                            or ''
                        }}"
                    >
                </div>

            </div>

            <p class="space">
                <label>
                    <input
                        style="width:auto"
                        type="checkbox"
                        name="
                            needs_relation_proof
                        "
                        {% if
                            student
                            .needs_relation_proof
                        %}
                            checked
                        {% endif %}
                    >

                    Нужен документ
                    о смене фамилии,
                    опеке или усыновлении
                </label>
            </p>

        </section>

        <div
            style="
                display:flex;
                gap:10px;
                flex-wrap:wrap;
            "
        >

            <button
                class="btn btn-primary"
                type="submit"
            >
                Сохранить изменения
            </button>

            <a
                class="btn btn-secondary"
                href="{{ url_for(
                    'student_detail',
                    student_id=student.id
                ) }}"
            >
                Отмена
            </a>

        </div>

    </form>
    """

    return render_page(
        "Редактирование данных",
        body,
        student=student,
    )


@app.route("/students/<int:student_id>")
@login_required
def student_detail(student_id: int):
    student = get_student_or_404(student_id)
    checklist = build_document_checklist(student)

    comments = get_db().execute(
        """
        SELECT comments.*, users.full_name AS author_name
        FROM comments
        JOIN users ON users.id = comments.created_by
        WHERE comments.student_id = ?
        ORDER BY comments.is_open DESC, comments.created_at DESC
        """,
        (student_id,),
    ).fetchall()

    required_count = sum(1 for item in checklist if item["required"])
    uploaded_required = sum(
        1
        for item in checklist
        if item["required"] and item["document"]
    )

    body = """
    <div style="display:flex;justify-content:space-between;gap:20px;align-items:center">
        <div>
            <h1>
                {{ student.last_name }}
                {{ student.first_name }}
                {{ student.middle_name or '' }}
            </h1>

            <p class="muted">
                {{ student.branch_name }} ·
                {{ student.class_number }} класс ·
                {{ calculate_age(student.birth_date) }} лет
            </p>
        </div>

        <span class="status {{ status_classes[student.status] }}">
            {{ status_labels[student.status] }}
        </span>
    </div>

    <div class="grid metrics">
        <div class="card metric-yellow">
            Комплектность
            <div class="metric-number">
                {{ uploaded_required }} / {{ required_count }}
            </div>
        </div>

        <div class="card">
            Открытые замечания
            <div class="metric-number">
                {{ comments | selectattr('is_open') | list | length }}
            </div>
        </div>

        <div class="card">
            Проверяющий
            <div style="font-size:18px;font-weight:700;margin-top:12px">
                {{ student.assigned_name or 'Не назначен' }}
            </div>
        </div>
    </div>

    
    {# NAME EDIT V9.2.1 BUTTON #}

    {% if current_user.role
        in ('branch', 'attestation', 'admin') %}

        <div
            class="space"
            style="
                display:flex;
                gap:10px;
                flex-wrap:wrap;
                margin-bottom:18px;
            "
        >

            <a
                class="btn btn-secondary"
                href="{{ url_for(
                    'edit_student_names_v9_2_1',
                    student_id=student.id
                ) }}"
            >
                ✎ Редактировать данные
            </a>

        </div>

    {% endif %}

<h2>Документы</h2>

    {% for group in ['Документы ребенка',
                     'Документы законного представителя',
                     'Заявления и согласия'] %}
        <div class="card space">
            <h3>{{ group }}</h3>

            {% for item in checklist if item.group == group %}
                <div class="document-row">
                    <div>
                        <strong>{{ item.name }}</strong><br>

                        {% if item.required %}
                            <span class="required">Обязательно</span>
                        {% else %}
                            <span class="optional">Не требуется</span>
                        {% endif %}

                        {% if item.help %}
                            <div class="muted">{{ item.help }}</div>
                        {% endif %}
                    </div>

                    <div>
                        {% if item.document %}
                            <span class="status green">Загружен</span>
                            <div class="muted">
                                версия {{ item.document.version }}
                            </div>
                        {% else %}
                            <span class="status gray">Нет файла</span>
                        {% endif %}
                    </div>

                    <div>
                        {% if item.document %}
                            <a
                                class="btn btn-secondary btn-small"
                                target="_blank"
                                href="{{ url_for(
                                    'download_document',
                                    document_id=item.document.id
                                ) }}"
                            >
                                Открыть
                            </a>
                        {% endif %}
                    </div>

                    <div>
                        {% if current_user.role == 'branch'
                              and student.status in ('draft', 'correction') %}
                            <a
                                class="btn btn-primary btn-small"
                                href="{{ url_for(
                                    'upload_document',
                                    student_id=student.id,
                                    document_type=item.code
                                ) }}"
                            >
                                {% if item.document %}Заменить{% else %}Загрузить{% endif %}
                            </a>
                        {% endif %}
                    </div>
                </div>
            {% endfor %}
        </div>
    {% endfor %}

    {% if comments %}
        <h2>Замечания</h2>

        <div class="card">
            {% for comment in comments %}
                <div style="padding:14px 0;border-bottom:1px solid #eee">
                    <strong>{{ comment.category }}</strong>

                    {% if comment.document_type %}
                        · {{ document_types[comment.document_type].name }}
                    {% endif %}

                    {% if comment.is_open %}
                        <span class="status red">Открыто</span>
                    {% else %}
                        <span class="status green">Исправлено</span>
                    {% endif %}

                    <p>{{ comment.text }}</p>

                    <div class="muted">
                        {{ comment.author_name }} ·
                        {{ comment.created_at[:16].replace('T', ' ') }}
                    </div>

                    {% if current_user.role == 'branch'
                          and comment.is_open
                          and student.status == 'correction' %}
                        <form
                            class="inline"
                            method="post"
                            action="{{ url_for(
                                'resolve_comment',
                                comment_id=comment.id
                            ) }}"
                        >
                            <input
                                type="hidden"
                                name="csrf_token"
                                value="{{ csrf_token }}"
                            >

                            <button class="btn btn-primary btn-small space">
                                Отметить исправленным
                            </button>
                        </form>
                    {% endif %}
                </div>
            {% endfor %}
        </div>
    {% endif %}

    <div class="space">
        {% if current_user.role == 'branch'
              and student.status in ('draft', 'correction') %}
            <form
                class="inline"
                method="post"
                action="{{ url_for('submit_student', student_id=student.id) }}"
            >
                <input
                    type="hidden"
                    name="csrf_token"
                    value="{{ csrf_token }}"
                >

                <button class="btn btn-primary">
                    Отправить в отдел аттестации
                </button>
            </form>
        {% endif %}

        {% if current_user.role in ('attestation', 'admin') %}
            <a
                class="btn btn-secondary"
                href="{{ url_for('review_student', student_id=student.id) }}"
            >
                Открыть проверку
            </a>
        {% endif %}
    </div>
    """

    return render_page(
        "Карточка ученика",
        body,
        student=student,
        checklist=checklist,
        comments=comments,
        calculate_age=calculate_age,
        uploaded_required=uploaded_required,
        required_count=required_count,
        document_types=DOCUMENT_TYPES,
    )


# ============================================================
# Документы
# ============================================================

@app.route(
    "/students/<int:student_id>/upload/<document_type>",
    methods=["GET", "POST"],
)
@roles_required("branch")
def upload_document(student_id: int, document_type: str):
    student = get_student_or_404(student_id)

    if student["status"] not in ("draft", "correction"):
        abort(400, "Карточка заблокирована для редактирования.")

    if document_type not in DOCUMENT_TYPES:
        abort(404)

    existing_documents = get_db().execute(
        """
        SELECT documents.*, document_type
        FROM documents
        WHERE student_id = ?
        ORDER BY uploaded_at DESC
        """,
        (student_id,),
    ).fetchall()

    if request.method == "POST":
        upload = request.files.get("file")
        reuse_document_id = request.form.get("reuse_document_id", "").strip()

        original_name = ""
        stored_name = ""
        mime_type = ""
        source_document_id = None

        if reuse_document_id:
            source = get_db().execute(
                """
                SELECT *
                FROM documents
                WHERE id = ? AND student_id = ?
                """,
                (int(reuse_document_id), student_id),
            ).fetchone()

            if not source:
                abort(400, "Исходный документ не найден.")

            original_name = source["original_name"]
            stored_name = source["stored_name"]
            mime_type = source["mime_type"]
            source_document_id = source["id"]

        elif upload and upload.filename:
            filename = secure_filename(upload.filename)

            if not allowed_file(filename):
                flash(
                    "Допустимы только PDF, JPG, JPEG и PNG.",
                    "error",
                )
                return redirect(request.url)

            extension = filename.rsplit(".", 1)[1].lower()
            stored_name = f"{uuid.uuid4().hex}.{extension}"
            original_name = filename
            mime_type = upload.mimetype

            upload.save(UPLOAD_DIR / stored_name)

        else:
            flash("Выберите файл или ранее загруженный документ.", "error")
            return redirect(request.url)

        latest_version = get_db().execute(
            """
            SELECT COALESCE(MAX(version), 0)
            FROM documents
            WHERE student_id = ? AND document_type = ?
            """,
            (student_id, document_type),
        ).fetchone()[0]

        get_db().execute(
            """
            INSERT INTO documents (
                student_id,
                document_type,
                original_name,
                stored_name,
                mime_type,
                version,
                source_document_id,
                uploaded_by,
                uploaded_at,
                ocr_status
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'not_configured')
            """,
            (
                student_id,
                document_type,
                original_name,
                stored_name,
                mime_type,
                latest_version + 1,
                source_document_id,
                g.current_user["id"],
                datetime.now().isoformat(timespec="seconds"),
            ),
        )

        get_db().execute(
            """
            UPDATE students
            SET updated_at = ?
            WHERE id = ?
            """,
            (
                datetime.now().isoformat(timespec="seconds"),
                student_id,
            ),
        )

        get_db().commit()

        audit(
            "document_uploaded",
            student_id,
            f"{document_type}; версия {latest_version + 1}",
        )

        flash("Документ загружен.", "success")
        return redirect(
            url_for("student_detail", student_id=student_id)
        )

    body = """
    <h1>{{ document_config.name }}</h1>

    <div class="alert info">
        Старые версии документов не удаляются.
        При замене система сохраняет историю.
    </div>

    <form class="form-section" method="post" enctype="multipart/form-data">
        <input type="hidden" name="csrf_token" value="{{ csrf_token }}">

        <div>
            <label>Новый файл</label>
            <input
                type="file"
                name="file"
                accept=".pdf,.jpg,.jpeg,.png"
            >
            <p class="muted">PDF, JPG, JPEG или PNG. Не более 25 МБ.</p>
        </div>

        {% if existing_documents %}
            <h3>Или использовать ранее загруженный файл</h3>

            <select name="reuse_document_id">
                <option value="">Не использовать</option>

                {% for document in existing_documents %}
                    <option value="{{ document.id }}">
                        {{ document_types[document.document_type].name }}
                        — {{ document.original_name }}
                        — версия {{ document.version }}
                    </option>
                {% endfor %}
            </select>
        {% endif %}

        <div
            class="space"
            style="
                display:flex;
                gap:12px;
                flex-wrap:wrap;
                align-items:center;
            "
        >
            <button
                class="btn btn-primary"
                type="submit"
            >
                Сохранить документ
            </button>

            <a
                class="btn btn-secondary"
                href="{{ url_for(
                    'student_detail',
                    student_id=student.id
                ) }}"
            >
                Назад
            </a>
        </div>
    </form>
    """

    return render_page(
        "Загрузка документа",
        body,
        student=student,
        document_config=DOCUMENT_TYPES[document_type],
        existing_documents=existing_documents,
        document_types=DOCUMENT_TYPES,
    )


@app.route("/documents/<int:document_id>")
@login_required
def download_document(document_id: int):
    document = get_db().execute(
        """
        SELECT documents.*, students.branch_id
        FROM documents
        JOIN students ON students.id = documents.student_id
        WHERE documents.id = ?
        """,
        (document_id,),
    ).fetchone()

    if not document:
        abort(404)

    if (
        g.current_user["role"] == "branch"
        and document["branch_id"] != g.current_user["branch_id"]
    ):
        abort(403)

    document_path = UPLOAD_DIR / document["stored_name"]

    if not document_path.is_file():
        abort(404, "Файл документа не найден в хранилище.")

    with document_path.open("rb") as file:
        signature = file.read(1024)

    if b"%PDF" in signature:
        document_mimetype = "application/pdf"
        document_extension = ".pdf"

    elif signature.startswith(b"\xff\xd8\xff"):
        document_mimetype = "image/jpeg"
        document_extension = ".jpg"

    elif signature.startswith(b"\x89PNG\r\n\x1a\n"):
        document_mimetype = "image/png"
        document_extension = ".png"

    else:
        stored_extension = Path(
            document["stored_name"]
        ).suffix.lower()

        if stored_extension in {".jpg", ".jpeg"}:
            document_mimetype = "image/jpeg"
            document_extension = ".jpg"

        elif stored_extension == ".png":
            document_mimetype = "image/png"
            document_extension = ".png"

        elif stored_extension == ".pdf":
            document_mimetype = "application/pdf"
            document_extension = ".pdf"

        else:
            document_mimetype = (
                document["mime_type"]
                or "application/octet-stream"
            )
            document_extension = stored_extension or ".bin"

    document_display_name = (
        f"document_{document_id}{document_extension}"
    )

    audit(
        "document_viewed",
        document["student_id"],
        (
            f"document_id={document_id}; "
            f"mimetype={document_mimetype}"
        ),
    )

    response = send_file(
        document_path,
        mimetype=document_mimetype,
        as_attachment=False,
        download_name=document_display_name,
        conditional=True,
        etag=True,
    )

    response.headers["Content-Disposition"] = (
        f'inline; filename="{document_display_name}"'
    )

    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["Cache-Control"] = (
        "private, no-store, max-age=0"
    )

    return response


# ============================================================
# Отправка и замечания
# ============================================================

@app.route("/students/<int:student_id>/submit", methods=["POST"])
@roles_required("branch")
def submit_student(student_id: int):
    student = get_student_or_404(student_id)

    if student["status"] not in ("draft", "correction"):
        abort(400)

    missing = missing_required_documents(student)

    open_comments = get_db().execute(
        """
        SELECT COUNT(*)
        FROM comments
        WHERE student_id = ? AND is_open = 1
        """,
        (student_id,),
    ).fetchone()[0]

    if missing:
        flash(
            "Нельзя отправить комплект. Не загружены: "
            + ", ".join(missing),
            "error",
        )
        return redirect(
            url_for("student_detail", student_id=student_id)
        )

    if open_comments:
        flash(
            "Сначала отметьте все замечания как исправленные.",
            "error",
        )
        return redirect(
            url_for("student_detail", student_id=student_id)
        )

    get_db().execute(
        """
        UPDATE students
        SET
            status = 'submitted',
            assigned_to = NULL,
            updated_at = ?
        WHERE id = ?
        """,
        (
            datetime.now().isoformat(timespec="seconds"),
            student_id,
        ),
    )
    get_db().commit()

    audit("student_submitted", student_id)

    flash("Карточка передана в отдел аттестации.", "success")
    return redirect(url_for("student_detail", student_id=student_id))


@app.route("/comments/<int:comment_id>/resolve", methods=["POST"])
@roles_required("branch")
def resolve_comment(comment_id: int):
    comment = get_db().execute(
        """
        SELECT comments.*, students.branch_id
        FROM comments
        JOIN students ON students.id = comments.student_id
        WHERE comments.id = ?
        """,
        (comment_id,),
    ).fetchone()

    if not comment:
        abort(404)

    if comment["branch_id"] != g.current_user["branch_id"]:
        abort(403)

    get_db().execute(
        """
        UPDATE comments
        SET
            is_open = 0,
            resolved_by = ?,
            resolved_at = ?
        WHERE id = ?
        """,
        (
            g.current_user["id"],
            datetime.now().isoformat(timespec="seconds"),
            comment_id,
        ),
    )
    get_db().commit()

    audit(
        "comment_marked_fixed",
        comment["student_id"],
        f"comment_id={comment_id}",
    )

    flash("Замечание отмечено как исправленное.", "success")

    return redirect(
        url_for(
            "student_detail",
            student_id=comment["student_id"],
        )
    )


# ============================================================
# Отдел аттестации
# ============================================================

@app.route("/review")
@roles_required("attestation", "admin")
def review_queue():
    mode = request.args.get("mode", "queue")

    query = """
        SELECT
            students.*,
            branches.name AS branch_name,
            users.full_name AS assigned_name,
            (
                SELECT COUNT(*)
                FROM comments
                WHERE comments.student_id = students.id
                  AND comments.is_open = 1
            ) AS open_comments
        FROM students
        JOIN branches ON branches.id = students.branch_id
        LEFT JOIN users ON users.id = students.assigned_to
        WHERE 1 = 1
    """

    params: list[Any] = []

    if mode == "mine":
        query += " AND students.assigned_to = ?"
        params.append(g.current_user["id"])
    else:
        query += """
            AND students.status IN (
                'submitted',
                'in_review',
                'correction'
            )
        """

    query += """
        ORDER BY
            CASE students.status
                WHEN 'submitted' THEN 1
                WHEN 'in_review' THEN 2
                WHEN 'correction' THEN 3
                ELSE 4
            END,
            students.updated_at ASC
    """

    rows = get_db().execute(query, params).fetchall()

    body = """
    <h1>Очередь проверки</h1>

    <div class="tabs">
        <a class="tab" href="{{ url_for('review_queue', mode='queue') }}">
            Общая очередь
        </a>

        <a class="tab" href="{{ url_for('review_queue', mode='mine') }}">
            Мои карточки
        </a>
    </div>

    <div class="card">
        <table>
            <thead>
                <tr>
                    <th>Ученик</th>
                    <th>Филиал</th>
                    <th>Класс</th>
                    <th>Статус</th>
                    <th>Проверяющий</th>
                    <th>Замечания</th>
                    <th></th>
                </tr>
            </thead>
            <tbody>
                {% for student in rows %}
                    <tr>
                        <td>
                            {{ student.last_name }}
                            {{ student.first_name }}
                            {{ student.middle_name or '' }}
                        </td>

                        <td>{{ student.branch_name }}</td>
                        <td>{{ student.class_number }}</td>

                        <td>
                            <span class="status {{ status_classes[student.status] }}">
                                {{ status_labels[student.status] }}
                            </span>
                        </td>

                        <td>{{ student.assigned_name or 'Не назначен' }}</td>
                        <td>{{ student.open_comments }}</td>

                        <td>
                            <a
                                class="btn btn-primary btn-small"
                                href="{{ url_for(
                                    'review_student',
                                    student_id=student.id
                                ) }}"
                            >
                                Открыть
                            </a>
                        </td>
                    </tr>
                {% else %}
                    <tr><td colspan="7">Очередь пуста.</td></tr>
                {% endfor %}
            </tbody>
        </table>
    </div>
    """

    return render_page("Очередь проверки", body, rows=rows)


@app.route("/review/<int:student_id>", methods=["GET", "POST"])
@roles_required("attestation", "admin")
def review_student(student_id: int):
    student = get_student_or_404(student_id)

    if request.method == "POST":
        action = request.form.get("action")

        if action == "take":
            if student["status"] != "submitted":
                flash("Карточка уже недоступна для взятия в работу.", "error")
            else:
                get_db().execute(
                    """
                    UPDATE students
                    SET
                        status = 'in_review',
                        assigned_to = ?,
                        updated_at = ?
                    WHERE id = ?
                    """,
                    (
                        g.current_user["id"],
                        datetime.now().isoformat(timespec="seconds"),
                        student_id,
                    ),
                )
                get_db().commit()

                audit("student_taken_for_review", student_id)
                flash("Карточка закреплена за вами.", "success")

            return redirect(
                url_for("review_student", student_id=student_id)
            )

        if action == "comment":
            category = request.form.get("category", "").strip()
            text = request.form.get("text", "").strip()
            document_type = request.form.get("document_type") or None

            if category not in COMMENT_CATEGORIES:
                flash("Выберите категорию замечания.", "error")
            else:
                get_db().execute(
                    """
                    INSERT INTO comments (
                        student_id,
                        document_type,
                        category,
                        text,
                        is_open,
                        created_by,
                        created_at
                    )
                    VALUES (?, ?, ?, ?, 1, ?, ?)
                    """,
                    (
                        student_id,
                        document_type,
                        category,
                        text,
                        g.current_user["id"],
                        datetime.now().isoformat(timespec="seconds"),
                    ),
                )
                get_db().commit()

                audit(
                    "comment_created",
                    student_id,
                    f"{category}: {text}",
                )

                flash("Замечание добавлено.", "success")

            return redirect(
                url_for("review_student", student_id=student_id)
            )

        if action == "return":
            open_comments = get_db().execute(
                """
                SELECT COUNT(*)
                FROM comments
                WHERE student_id = ? AND is_open = 1
                """,
                (student_id,),
            ).fetchone()[0]

            if not open_comments:
                flash(
                    "Перед возвратом добавьте хотя бы одно замечание.",
                    "error",
                )
            else:
                get_db().execute(
                    """
                    UPDATE students
                    SET
                        status = 'correction',
                        assigned_to = NULL,
                        updated_at = ?
                    WHERE id = ?
                    """,
                    (
                        datetime.now().isoformat(timespec="seconds"),
                        student_id,
                    ),
                )
                get_db().commit()

                audit("student_returned", student_id)
                flash("Карточка возвращена филиалу.", "success")

                return redirect(url_for("review_queue"))

        if action == "ready":
            missing = missing_required_documents(student)

            open_comments = get_db().execute(
                """
                SELECT COUNT(*)
                FROM comments
                WHERE student_id = ? AND is_open = 1
                """,
                (student_id,),
            ).fetchone()[0]

            if missing:
                flash(
                    "Не загружены обязательные документы: "
                    + ", ".join(missing),
                    "error",
                )
            elif open_comments:
                flash("Есть открытые замечания.", "error")
            else:
                get_db().execute(
                    """
                    UPDATE students
                    SET
                        status = 'ready',
                        updated_at = ?
                    WHERE id = ?
                    """,
                    (
                        datetime.now().isoformat(timespec="seconds"),
                        student_id,
                    ),
                )
                get_db().commit()

                audit("student_ready", student_id)
                flash("Карточка готова к зачислению.", "success")

                return redirect(url_for("review_queue"))

        if action == "enroll":
            if student["status"] != "ready":
                flash(
                    "Зачислить можно только карточку в статусе "
                    "«Готово к зачислению».",
                    "error",
                )
            else:
                get_db().execute(
                    """
                    UPDATE students
                    SET
                        status = 'enrolled',
                        updated_at = ?
                    WHERE id = ?
                    """,
                    (
                        datetime.now().isoformat(timespec="seconds"),
                        student_id,
                    ),
                )
                get_db().commit()

                audit("student_enrolled", student_id)
                flash("Ученик отмечен как зачисленный.", "success")

            return redirect(
                url_for("review_student", student_id=student_id)
            )

    student = get_student_or_404(student_id)
    checklist = build_document_checklist(student)

    comments = get_db().execute(
        """
        SELECT comments.*, users.full_name AS author_name
        FROM comments
        JOIN users ON users.id = comments.created_by
        WHERE comments.student_id = ?
        ORDER BY comments.created_at DESC
        """,
        (student_id,),
    ).fetchall()

    body = """
    <div style="display:flex;justify-content:space-between;align-items:center">
        <div>
            <h1>
                Проверка:
                {{ student.last_name }}
                {{ student.first_name }}
            </h1>

            <p class="muted">
                {{ student.branch_name }} ·
                {{ student.class_number }} класс
            </p>
        </div>

        <span class="status {{ status_classes[student.status] }}">
            {{ status_labels[student.status] }}
        </span>
    </div>

    {% if student.status == 'submitted' %}
        <form method="post">
            <input type="hidden" name="csrf_token" value="{{ csrf_token }}">
            <input type="hidden" name="action" value="take">

            <button class="btn btn-primary">
                Взять в работу
            </button>
        </form>
    {% endif %}

    <div class="form-grid space">
        <div class="card">
            <h2>Ученик</h2>
            <p><strong>ФИО:</strong> {{ student_full_name }}</p>
            <p><strong>Дата рождения:</strong> {{ student.birth_date }}</p>
            <p><strong>СНИЛС:</strong> {{ student.student_snils or '—' }}</p>
            <p>
                <strong>Свидетельство:</strong>
                {{ student.birth_certificate_series or '' }}
                {{ student.birth_certificate_number or '—' }}
            </p>
            <p>
                <strong>Паспорт:</strong>
                {{ student.child_passport_series or '' }}
                {{ student.child_passport_number or '—' }}
            </p>
        </div>

        <div class="card">
            <h2>Законный представитель</h2>
            <p><strong>ФИО:</strong> {{ parent_full_name }}</p>
            <p><strong>Связь:</strong> {{ student.relation_type }}</p>
            <p><strong>СНИЛС:</strong> {{ student.parent_snils or '—' }}</p>
            <p>
                <strong>Паспорт:</strong>
                {{ student.parent_passport_series or '' }}
                {{ student.parent_passport_number or '—' }}
            </p>
        </div>
    </div>

    <h2>Документы</h2>

    <div class="card">
        {% for item in checklist %}
            <div class="document-row">
                <div>
                    <strong>{{ item.name }}</strong>

                    {% if item.required %}
                        <span class="required"> *</span>
                    {% endif %}
                </div>

                <div>
                    {% if item.document %}
                        <span class="status green">Есть</span>
                    {% else %}
                        <span class="status red">Нет</span>
                    {% endif %}
                </div>

                <div>
                    {% if item.document %}
                        <a
                            class="btn btn-secondary btn-small"
                            target="_blank"
                            href="{{ url_for(
                                'download_document',
                                document_id=item.document.id
                            ) }}"
                        >
                            Открыть
                        </a>
                    {% endif %}
                </div>

                <div>
                    {% if item.document %}
                        {{ item.document.original_name }}
                    {% endif %}
                </div>
            </div>
        {% endfor %}
    </div>

    <h2>Добавить замечание</h2>

    <form class="card" method="post">
        <input type="hidden" name="csrf_token" value="{{ csrf_token }}">
        <input type="hidden" name="action" value="comment">

        <div class="form-grid">
            <div>
                <label>Документ</label>
                <select name="document_type">
                    <option value="">Общее замечание</option>

                    {% for code, config in document_types.items() %}
                        <option value="{{ code }}">{{ config.name }}</option>
                    {% endfor %}
                </select>
            </div>

            <div>
                <label>Категория</label>
                <select name="category" required>
                    {% for category in categories %}
                        <option>{{ category }}</option>
                    {% endfor %}
                </select>
            </div>
        </div>

        <div class="space">
            <label>Комментарий (необязательно)</label>
            <textarea
                name="text"
                placeholder="При необходимости добавьте пояснение"
            ></textarea>
        </div>

        <button class="btn btn-primary space">
            Добавить замечание
        </button>
    </form>

    {% if comments %}
        <h2>История замечаний</h2>

        <div class="card">
            {% for comment in comments %}
                <div style="padding:12px 0;border-bottom:1px solid #eee">
                    <strong>{{ comment.category }}</strong>

                    {% if comment.is_open %}
                        <span class="status red">Открыто</span>
                    {% else %}
                        <span class="status green">Исправлено</span>
                    {% endif %}

                    <p>{{ comment.text }}</p>
                    <span class="muted">{{ comment.author_name }}</span>
                </div>
            {% endfor %}
        </div>
    {% endif %}

    <div class="space">
        {% if student.status == 'in_review' %}
            <form class="inline" method="post">
                <input type="hidden" name="csrf_token" value="{{ csrf_token }}">
                <input type="hidden" name="action" value="return">

                <button class="btn btn-danger">
                    Вернуть в филиал
                </button>
            </form>

            <form class="inline" method="post">
                <input type="hidden" name="csrf_token" value="{{ csrf_token }}">
                <input type="hidden" name="action" value="ready">

                <button class="btn btn-green">
                    Готово к зачислению
                </button>
            </form>
        {% endif %}

        {% if student.status == 'ready' %}
            <form class="inline" method="post">
                <input type="hidden" name="csrf_token" value="{{ csrf_token }}">
                <input type="hidden" name="action" value="enroll">

                <button class="btn btn-green">
                    Зачислен
                </button>
            </form>
        {% endif %}
    </div>
    """

    return render_page(
        "Проверка карточки",
        body,
        student=student,
        checklist=checklist,
        comments=comments,
        document_types=DOCUMENT_TYPES,
        categories=COMMENT_CATEGORIES,
        student_full_name=full_name(
            student["last_name"],
            student["first_name"],
            student["middle_name"],
        ),
        parent_full_name=full_name(
            student["parent_last_name"],
            student["parent_first_name"],
            student["parent_middle_name"],
        ),
    )


# ============================================================
# Выгрузка Дневник.ру
# ============================================================

@app.route("/export")
@roles_required("attestation", "admin")
def export_page():
    branches = get_db().execute(
        "SELECT * FROM branches ORDER BY name"
    ).fetchall()

    body = """
    <h1>Выгрузка для Дневник.ру</h1>

    <div class="alert info">
        В один файл включаются ученики и представители одного класса.
        В выгрузку попадают только карточки со статусом
        «Готово к зачислению».
    </div>

    <form
        class="form-section"
        method="get"
        action="{{ url_for('export_xlsx') }}"
    >
        <div class="form-grid">
            <div>
                <label>Филиал</label>
                <select name="branch_id" required>
                    {% for branch in branches %}
                        <option value="{{ branch.id }}">
                            {{ branch.name }}
                        </option>
                    {% endfor %}
                </select>
            </div>

            <div>
                <label>Класс</label>
                <input
                    type="number"
                    name="class_number"
                    min="1"
                    max="11"
                    required
                >
            </div>
        </div>

        <button class="btn btn-primary space">
            Сформировать Excel
        </button>
    </form>
    """

    return render_page(
        "Выгрузка Дневник.ру",
        body,
        branches=branches,
    )


@app.route("/export.xlsx")
@roles_required("attestation", "admin")
def export_xlsx():
    branch_id = request.args.get("branch_id", type=int)
    class_number = request.args.get("class_number", type=int)

    if not branch_id or not class_number:
        abort(400)

    branch = get_db().execute(
        "SELECT * FROM branches WHERE id = ?",
        (branch_id,),
    ).fetchone()

    if not branch:
        abort(404)

    students_rows = get_db().execute(
        """
        SELECT *
        FROM students
        WHERE branch_id = ?
          AND class_number = ?
          AND status = 'ready'
        ORDER BY last_name, first_name
        """,
        (branch_id, class_number),
    ).fetchall()

    headers = [
        "ФИО законного представителя (целиком)",
        "Пол законного представителя",
        "Дата рождения законного представителя",
        "СНИЛС законного представителя",
        "Паспорт РФ законного представителя — серия",
        "Паспорт РФ законного представителя — номер",
        "Вид родственной связи",
        "ФИО обучающегося (целиком)",
        "Пол обучающегося",
        "Дата рождения обучающегося",
        "Дата прибытия обучающегося в ОО",
        "Дата прибытия обучающегося в класс",
        "СНИЛС обучающегося",
        "Свидетельство о рождении — серия",
        "Свидетельство о рождении — номер",
        "Паспорт РФ обучающегося — серия",
        "Паспорт РФ обучающегося — номер",
        "Мобильный телефон законного представителя",
        "Email законного представителя",
        "Мобильный телефон обучающегося",
        "Email обучающегося",
    ]

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Импорт"

    worksheet.append(headers)

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="FFD500",
    )

    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    for student in students_rows:
        age = calculate_age(student["birth_date"])

        birth_series = (
            student["birth_certificate_series"]
            if age < 14
            else ""
        )
        birth_number = (
            student["birth_certificate_number"]
            if age < 14
            else ""
        )

        passport_series = (
            student["child_passport_series"]
            if age >= 14
            else ""
        )
        passport_number = (
            student["child_passport_number"]
            if age >= 14
            else ""
        )

        worksheet.append(
            [
                full_name(
                    student["parent_last_name"],
                    student["parent_first_name"],
                    student["parent_middle_name"],
                ),
                student["parent_gender"] or "",
                student["parent_birth_date"] or "",
                student["parent_snils"] or "",
                student["parent_passport_series"] or "",
                student["parent_passport_number"] or "",
                student["relation_type"],
                full_name(
                    student["last_name"],
                    student["first_name"],
                    student["middle_name"],
                ),
                student["gender"] or "",
                student["birth_date"],
                student["arrival_school_date"] or "",
                student["arrival_class_date"] or "",
                student["student_snils"] or "",
                birth_series,
                birth_number,
                passport_series,
                passport_number,
                student["parent_phone"] or "",
                student["parent_email"] or "",
                student["student_phone"] or "",
                student["student_email"] or "",
            ]
        )

    for column in worksheet.columns:
        max_length = max(
            len(str(cell.value or ""))
            for cell in column
        )
        worksheet.column_dimensions[column[0].column_letter].width = min(
            max(max_length + 2, 14),
            45,
        )

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    audit(
        "xlsx_exported",
        details=(
            f"branch={branch['name']}; "
            f"class={class_number}; "
            f"rows={len(students_rows)}"
        ),
    )

    filename = (
        f"Дневник_{branch['name']}_"
        f"{class_number}_класс.xlsx"
    )

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
    )


# ============================================================
# Администратор
# ============================================================

@app.route("/admin/users", methods=["GET", "POST"])
@roles_required("admin")
def admin_users():
    # ADMIN USER TOGGLE V1
    if request.method == "POST":
        action = request.form.get("action", "").strip()

        # ---------------------------------------------
        # Включение / отключение пользователя
        # ---------------------------------------------
        if action == "toggle_user":
            user_id = request.form.get(
                "user_id",
                type=int,
            )

            if not user_id:
                flash(
                    "Не удалось определить пользователя.",
                    "error",
                )
                return redirect(
                    url_for("admin_users")
                )

            user = get_db().execute(
                """
                SELECT *
                FROM users
                WHERE id = ?
                """,
                (user_id,),
            ).fetchone()

            if not user:
                flash(
                    "Пользователь не найден.",
                    "error",
                )
                return redirect(
                    url_for("admin_users")
                )

            # Текущий администратор не может
            # случайно заблокировать сам себя.
            if (
                g.current_user
                and user["id"]
                == g.current_user["id"]
            ):
                flash(
                    "Нельзя отключить собственную "
                    "учетную запись администратора.",
                    "error",
                )
                return redirect(
                    url_for("admin_users")
                )

            new_active = (
                0
                if user["active"]
                else 1
            )

            get_db().execute(
                """
                UPDATE users
                SET active = ?
                WHERE id = ?
                """,
                (
                    new_active,
                    user_id,
                ),
            )

            get_db().commit()

            if new_active:
                audit(
                    "user_activated",
                    details=user["email"],
                )
                flash(
                    f"Пользователь {user['email']} включен.",
                    "success",
                )
            else:
                audit(
                    "user_deactivated",
                    details=user["email"],
                )
                flash(
                    f"Пользователь {user['email']} отключен.",
                    "success",
                )

            return redirect(
                url_for("admin_users")
            )

        email = request.form.get("email", "").strip().lower()
        full_user_name = request.form.get("full_name", "").strip()
        role = request.form.get("role", "")
        branch_id = request.form.get("branch_id", type=int)
        password = request.form.get("password", "")

        if not email.endswith("@top-academy.ru"):
            flash(
                "Разрешены только адреса @top-academy.ru.",
                "error",
            )
        elif role not in ("branch", "attestation", "admin"):
            flash("Некорректная роль.", "error")
        elif role == "branch" and not branch_id:
            flash("Для филиального пользователя выберите филиал.", "error")
        elif not full_user_name or len(password) < 8:
            flash(
                "Укажите ФИО и пароль не короче 8 символов.",
                "error",
            )
        else:
            try:
                get_db().execute(
                    """
                    INSERT INTO users (
                        email,
                        password_hash,
                        full_name,
                        role,
                        branch_id,
                        active,
                        created_at
                    )
                    VALUES (?, ?, ?, ?, ?, 1, ?)
                    """,
                    (
                        email,
                        generate_password_hash(password),
                        full_user_name,
                        role,
                        branch_id if role == "branch" else None,
                        datetime.now().isoformat(timespec="seconds"),
                    ),
                )
                get_db().commit()

                audit("user_created", details=email)
                flash("Пользователь создан.", "success")

            except sqlite3.IntegrityError:
                flash("Такой email уже зарегистрирован.", "error")

    users = get_db().execute(
        """
        SELECT users.*, branches.name AS branch_name
        FROM users
        LEFT JOIN branches ON branches.id = users.branch_id
        ORDER BY users.full_name
        """
    ).fetchall()

    branches = get_db().execute(
        "SELECT * FROM branches ORDER BY name"
    ).fetchall()

    body = """
    <h1>Пользователи</h1>

    <form class="form-section" method="post">
        <input type="hidden" name="csrf_token" value="{{ csrf_token }}">

        <div class="form-grid">
            <div>
                <label>ФИО</label>
                <input name="full_name" required>
            </div>

            <div>
                <label>Email</label>
                <input
                    type="email"
                    name="email"
                    placeholder="name@top-academy.ru"
                    required
                >
            </div>

            <div>
                <label>Роль</label>
                <select name="role" required>
                    <option value="branch">Филиал</option>
                    <option value="attestation">Отдел аттестации</option>
                    <option value="admin">Администратор</option>
                </select>
            </div>

            <div>
                <label>Филиал</label>
                <select name="branch_id">
                    <option value="">Не выбран</option>

                    {% for branch in branches %}
                        <option value="{{ branch.id }}">
                            {{ branch.name }}
                        </option>
                    {% endfor %}
                </select>
            </div>

            <div>
                <label>Временный пароль</label>
                <input type="password" name="password" required>
            </div>
        </div>

        <button class="btn btn-primary space">
            Добавить пользователя
        </button>
    </form>

    <div class="card">
        <table>
            <thead>
                <tr>
                    <th>ФИО</th>
                    <th>Email</th>
                    <th>Роль</th>
                    <th>Филиал</th>
                    <th>Статус</th>
                    <th>Действие</th>
                </tr>
            </thead>
            <tbody>
                {% for user in users %}
                    <tr>
                        <td>{{ user.full_name }}</td>
                        <td>{{ user.email }}</td>
                        <td>{{ user.role }}</td>
                        <td>{{ user.branch_name or '—' }}</td>

                        <td>
                            {% if user.active %}
                                <span
                                    style="
                                        display:inline-block;
                                        padding:5px 10px;
                                        border-radius:999px;
                                        background:#e8f7ec;
                                        color:#176b2c;
                                        font-weight:600;
                                    "
                                >
                                    Активен
                                </span>
                            {% else %}
                                <span
                                    style="
                                        display:inline-block;
                                        padding:5px 10px;
                                        border-radius:999px;
                                        background:#f1f1f1;
                                        color:#666;
                                        font-weight:600;
                                    "
                                >
                                    Отключен
                                </span>
                            {% endif %}
                        </td>

                        <td>
                            {% if
                                current_user
                                and
                                user.id == current_user.id
                            %}
                                <span class="muted">
                                    Текущий пользователь
                                </span>

                            {% else %}

                                <form
                                    method="post"
                                    style="margin:0"
                                    onsubmit="
                                        return confirm(
                                            '{{ 'Отключить' if user.active else 'Включить' }} пользователя {{ user.email }}?'
                                        );
                                    "
                                >
                                    <input
                                        type="hidden"
                                        name="csrf_token"
                                        value="{{ csrf_token }}"
                                    >

                                    <input
                                        type="hidden"
                                        name="action"
                                        value="toggle_user"
                                    >

                                    <input
                                        type="hidden"
                                        name="user_id"
                                        value="{{ user.id }}"
                                    >

                                    {% if user.active %}
                                        <button
                                            type="submit"
                                            class="btn btn-secondary"
                                            style="
                                                white-space:nowrap;
                                                border-color:#c94b4b;
                                                color:#a32d2d;
                                            "
                                        >
                                            Отключить
                                        </button>
                                    {% else %}
                                        <button
                                            type="submit"
                                            class="btn btn-secondary"
                                            style="
                                                white-space:nowrap;
                                            "
                                        >
                                            Включить
                                        </button>
                                    {% endif %}
                                </form>

                            {% endif %}
                        </td>
                    </tr>
                {% endfor %}
            </tbody>
        </table>
    </div>
    """

    return render_page(
        "Пользователи",
        body,
        users=users,
        branches=branches,
    )


# ============================================================
# Ошибки
# ============================================================

@app.errorhandler(403)
def forbidden(_: Any):
    return render_page(
        "Нет доступа",
        """
        <div class="card">
            <h1>Нет доступа</h1>
            <p>У вас нет прав для просмотра этого раздела или карточки.</p>
        </div>
        """,
    ), 403


@app.errorhandler(404)
def not_found(_: Any):
    return render_page(
        "Не найдено",
        """
        <div class="card">
            <h1>Страница не найдена</h1>
        </div>
        """,
    ), 404


@app.errorhandler(413)
def too_large(_: Any):
    return render_page(
        "Файл слишком большой",
        """
        <div class="card">
            <h1>Файл слишком большой</h1>
            <p>Максимальный размер одного файла — 25 МБ.</p>
        </div>
        """,
    ), 413


# ============================================================
# Запуск
# ============================================================


from portal_patch import apply_patch
apply_patch(app, globals())


from portal_validation_v6 import apply_validation_v6
apply_validation_v6(app, globals())


from portal_post_enrollment_v8 import apply_post_enrollment_v8
apply_post_enrollment_v8(app, globals())

from portal_order_sync_v9_4 import apply_order_sync_v9_4
apply_order_sync_v9_4(app, globals())


from portal_attestation_tools_v9 import apply_attestation_tools_v9
apply_attestation_tools_v9(app, globals())


from portal_post_docs_history_v9_5 import apply_post_docs_history_v9_5
apply_post_docs_history_v9_5(app, globals())


from portal_passport_ocr_v6_7 import apply_passport_ocr_v6_7
apply_passport_ocr_v6_7(app, globals())


from portal_students_tools_v9_7 import apply_students_tools_v9_7
apply_students_tools_v9_7(app, globals())


from portal_registration_ocr_v6_8 import apply_registration_ocr_v6_8
apply_registration_ocr_v6_8(app, globals())


from portal_post_docs_formats_v9_5_2 import apply_post_docs_formats_v9_5_2
apply_post_docs_formats_v9_5_2(app, globals())



from portal_post_docs_buttons_v9_5_4 import apply_post_docs_buttons_v9_5_4
apply_post_docs_buttons_v9_5_4(app, globals())


# MULTIFILE UPLOAD FIX V9.8
from portal_multifile_upload_fix_v9_8 import apply_multifile_upload_fix_v9_8
apply_multifile_upload_fix_v9_8(app, globals())


from portal_review_admin_v9_9 import apply_review_admin_v9_9
apply_review_admin_v9_9(app, globals())


# WORKFLOW FIX V9.9.1
from portal_workflow_fix_v9_9_1 import apply_workflow_fix_v9_9_1
apply_workflow_fix_v9_9_1(app, globals())


# DOCUMENT CONSISTENCY FIX V9.9.2
from portal_document_consistency_v9_9_2 import apply_document_consistency_v9_9_2
apply_document_consistency_v9_9_2(app, globals())


# LATEST DOCUMENT FIX V9.9.3
from portal_latest_document_v9_9_3 import apply_latest_document_v9_9_3
apply_latest_document_v9_9_3(app, globals())


# READY DOCUMENTS FIX V9.9.4
from portal_ready_documents_v9_9_4 import apply_ready_documents_v9_9_4
apply_ready_documents_v9_9_4(app, globals())

if __name__ == "__main__":
    init_db()

    app.run(
        host="127.0.0.1",
        port=5000,
        debug=True,
    )