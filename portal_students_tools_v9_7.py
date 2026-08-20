from __future__ import annotations

import io
from datetime import datetime
from typing import Any

from flask import (
    abort,
    g,
    request,
    send_file,
    url_for,
)

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter


def apply_students_tools_v9_7(
    app,
    namespace: dict[str, Any],
) -> None:

    if app.extensions.get(
        "students_tools_v9_7"
    ):
        return

    app.extensions[
        "students_tools_v9_7"
    ] = True

    get_db = namespace["get_db"]
    login_required = namespace["login_required"]
    render_page = namespace["render_page"]
    calculate_age = namespace["calculate_age"]
    full_name = namespace["full_name"]
    audit = namespace.get("audit")

    status_labels = namespace["STATUS_LABELS"]

    # ========================================================
    # ОБЩИЙ ЗАПРОС
    #
    # Используется одновременно:
    # - экраном Ученики;
    # - счетчиком;
    # - Excel.
    #
    # Поэтому список и выгрузка
    # всегда совпадают.
    # ========================================================

    # STUDENTS MULTI BRANCH V9.7.2

    # ========================================================
    # STUDENTS MULTI BRANCH V9.7.2
    # ========================================================

    def get_selected_branch_ids():

        result = []

        for value in request.args.getlist(
            "branch_id"
        ):

            try:
                branch_id = int(
                    value
                )

            except (
                TypeError,
                ValueError,
            ):
                continue

            if (
                branch_id > 0
                and branch_id not in result
            ):
                result.append(
                    branch_id
                )

        return result


    def build_students_query():

        query = """
            SELECT
                students.*,
                branches.name AS branch_name,

                (
                    SELECT COUNT(*)
                    FROM comments
                    WHERE
                        comments.student_id =
                            students.id
                        AND comments.is_open = 1
                ) AS open_comments

            FROM students

            JOIN branches
                ON branches.id =
                   students.branch_id

            WHERE 1 = 1
        """

        params = []

        role = g.current_user["role"]

        # --------------------------------
        # Филиал видит только себя.
        # --------------------------------

        if role == "branch":

            query += """
                AND students.branch_id = ?
            """

            params.append(
                g.current_user["branch_id"]
            )

        # --------------------------------
        # Филиал
        #
        # Для аттестации и администратора.
        # --------------------------------
        # Филиалы
        #
        # Аттестация / администратор
        # могут выбрать любое количество.
        # --------------------------------

        branch_ids = (
            get_selected_branch_ids()
        )

        if (
            role in (
                "attestation",
                "admin",
            )
            and branch_ids
        ):

            placeholders = ", ".join(
                "?"
                for _ in branch_ids
            )

            query += (
                " AND students.branch_id "
                f"IN ({placeholders}) "
            )

            params.extend(
                branch_ids
            )

        # --------------------------------
        # Поиск
        # --------------------------------

        search = request.args.get(
            "search",
            "",
        ).strip()

        if search:

            query += """
                AND (
                    students.last_name
                        LIKE ?

                    OR students.first_name
                        LIKE ?

                    OR students.middle_name
                        LIKE ?

                    OR (
                        students.last_name
                        || ' '
                        || students.first_name
                        || ' '
                        || coalesce(
                            students.middle_name,
                            ''
                        )
                    ) LIKE ?

                    OR students.student_snils
                        LIKE ?
                )
            """

            pattern = (
                f"%{search}%"
            )

            params.extend(
                [
                    pattern,
                    pattern,
                    pattern,
                    pattern,
                    pattern,
                ]
            )

        # --------------------------------
        # Статус
        #
        # Здесь НЕТ ограничения
        # по статусам.
        # Без фильтра выгружаются
        # абсолютно все статусы.
        # --------------------------------

        status = request.args.get(
            "status",
            "",
        ).strip()

        if status in status_labels:

            query += """
                AND students.status = ?
            """

            params.append(
                status
            )

        # --------------------------------
        # Класс
        # --------------------------------

        class_number = request.args.get(
            "class_number",
            "",
        ).strip()

        if class_number.isdigit():

            number = int(
                class_number
            )

            if 1 <= number <= 11:

                query += """
                    AND students.class_number = ?
                """

                params.append(
                    number
                )

        return query, params

    # ========================================================
    # СТРАНИЦА «УЧЕНИКИ»
    # ========================================================

    def students_v9_7():

        db = get_db()

        query, params = (
            build_students_query()
        )

        rows = db.execute(
            query
            + """
                ORDER BY
                    students.updated_at DESC
            """,
            params,
        ).fetchall()

        # Используем именно длину
        # итоговой выборки:
        # счетчик всегда соответствует
        # отображаемой таблице.
        total_count = len(
            rows
        )

        role = g.current_user[
            "role"
        ]

        selected_branch_ids = (
            get_selected_branch_ids()
        )

        # Сохраняем полный query string,
        # включая повторяющиеся branch_id,
        # чтобы Excel выгружал ровно
        # ту же выборку.
        export_url = url_for(
            "students_export_v9_7"
        )

        if request.query_string:

            export_url += (
                "?"
                + request.query_string.decode(
                    "utf-8"
                )
            )

        branches = []

        if role in (
            "attestation",
            "admin",
        ):

            branches = db.execute(
                """
                SELECT *
                FROM branches
                ORDER BY name
                """
            ).fetchall()

        body = """
        <div
            style="
                display:flex;
                justify-content:
                    space-between;
                align-items:center;
                gap:20px;
                flex-wrap:wrap;
            "
        >

            <h1>
                Ученики
            </h1>

            <div
                style="
                    display:flex;
                    gap:10px;
                    flex-wrap:wrap;
                "
            >

                <a
                    class="btn btn-secondary"
                    href="{{ export_url }}"
                >
                    Скачать Excel
                </a>

                {% if
                    current_user.role
                    == 'branch'
                %}

                    <a
                        class="
                            btn
                            btn-primary
                        "
                        href="{{
                            url_for(
                                'new_student'
                            )
                        }}"
                    >
                        + Добавить ученика
                    </a>

                {% endif %}

            </div>

        </div>


        <form
            class="card form-grid"
            method="get"
        >

            <div>

                <label>
                    Поиск
                </label>

                <input
                    name="search"
                    value="{{
                        request.args.get(
                            'search',
                            ''
                        )
                    }}"
                    placeholder="
                        ФИО или СНИЛС
                    "
                >

            </div>


            {% if
                current_user.role
                in (
                    'attestation',
                    'admin'
                )
            %}

                <div>

                    <label>
                        Филиал
                    </label>

                    
                    <details
                        style="
                            border:1px solid #ccc;
                            border-radius:10px;
                            background:#fff;
                            min-height:52px;
                        "
                    >

                        <summary
                            style="
                                cursor:pointer;
                                padding:14px 18px;
                                list-style:none;
                                user-select:none;
                            "
                        >

                            {% if
                                selected_branch_ids
                            %}

                                Выбрано филиалов:
                                {{
                                    selected_branch_ids
                                    | length
                                }}

                            {% else %}

                                Все филиалы

                            {% endif %}

                        </summary>


                        <div
                            style="
                                border-top:
                                    1px solid #e5e5e5;
                                padding:12px 16px;
                                max-height:300px;
                                overflow-y:auto;
                            "
                        >

                            {% for branch in branches %}

                                <label
                                    style="
                                        display:flex;
                                        align-items:center;
                                        gap:9px;
                                        padding:7px 2px;
                                        cursor:pointer;
                                        font-weight:400;
                                    "
                                >

                                    <input
                                        type="checkbox"
                                        name="branch_id"
                                        value="{{ branch.id }}"

                                        {% if
                                            branch.id
                                            in selected_branch_ids
                                        %}
                                            checked
                                        {% endif %}

                                        style="
                                            width:auto;
                                            margin:0;
                                        "
                                    >

                                    <span>
                                        {{ branch.name }}
                                    </span>

                                </label>

                            {% endfor %}

                        </div>

                    </details>


                </div>

            {% endif %}


            <div>

                <label>
                    Статус
                </label>

                <select
                    name="status"
                >

                    <option value="">
                        Все статусы
                    </option>

                    {% for
                        code,
                        label
                        in status_labels.items()
                    %}

                        <option
                            value="{{ code }}"

                            {% if
                                request.args.get(
                                    'status'
                                )
                                == code
                            %}
                                selected
                            {% endif %}
                        >
                            {{ label }}
                        </option>

                    {% endfor %}

                </select>

            </div>


            <div>

                <label>
                    Класс
                </label>

                <input
                    type="number"
                    min="1"
                    max="11"
                    name="class_number"
                    value="{{
                        request.args.get(
                            'class_number',
                            ''
                        )
                    }}"
                >

            </div>


            <div
                style="
                    align-self:end;
                    display:flex;
                    gap:8px;
                "
            >

                <button
                    class="
                        btn
                        btn-primary
                    "
                >
                    Применить
                </button>

                <a
                    class="
                        btn
                        btn-secondary
                    "
                    href="{{
                        url_for(
                            'students'
                        )
                    }}"
                >
                    Сбросить
                </a>

            </div>

        </form>


        {# ============================================== #}
        {# СЧЕТЧИК                                     #}
        {# Только аттестация + администратор            #}
        {# ============================================== #}

        {% if
            current_user.role
            in (
                'attestation',
                'admin'
            )
        %}

            <div
                style="
                    margin-top:22px;
                    margin-bottom:7px;
                    font-size:28px;
                    line-height:1;
                    font-weight:700;
                "
            >
                {{ total_count }}
            </div>

        {% endif %}


        <div
            class="card"
            style="
                margin-top:
                {% if
                    current_user.role
                    in (
                        'attestation',
                        'admin'
                    )
                %}
                    0
                {% else %}
                    18px
                {% endif %};
            "
        >

            <div
                style="
                    overflow-x:auto;
                "
            >

                <table>

                    <thead>

                        <tr>

                            <th>
                                Ученик
                            </th>

                            {% if
                                current_user.role
                                != 'branch'
                            %}

                                <th>
                                    Филиал
                                </th>

                            {% endif %}

                            <th>
                                Класс
                            </th>

                            <th>
                                Возраст
                            </th>

                            <th>
                                Статус
                            </th>

                            <th>
                                Замечания
                            </th>

                        </tr>

                    </thead>


                    <tbody>

                        {% for
                            student
                            in rows
                        %}

                            <tr>

                                <td>

                                    <a
                                        href="{{
                                            url_for(
                                                'student_detail',
                                                student_id=
                                                    student.id
                                            )
                                        }}"
                                    >
                                        {{
                                            student.last_name
                                        }}
                                        {{
                                            student.first_name
                                        }}
                                        {{
                                            student.middle_name
                                            or ''
                                        }}
                                    </a>

                                </td>


                                {% if
                                    current_user.role
                                    != 'branch'
                                %}

                                    <td>
                                        {{
                                            student
                                            .branch_name
                                        }}
                                    </td>

                                {% endif %}


                                <td>
                                    {{
                                        student
                                        .class_number
                                    }}
                                </td>


                                <td>
                                    {{
                                        calculate_age(
                                            student
                                            .birth_date
                                        )
                                    }}
                                </td>


                                <td>

                                    <span
                                        class="
                                            status
                                            {{
                                                status_classes[
                                                    student.status
                                                ]
                                            }}
                                        "
                                    >
                                        {{
                                            status_labels[
                                                student.status
                                            ]
                                        }}
                                    </span>

                                </td>


                                <td>
                                    {{
                                        student
                                        .open_comments
                                    }}
                                </td>

                            </tr>

                        {% else %}

                            <tr>

                                <td
                                    colspan="6"
                                >
                                    Ученики
                                    не найдены.
                                </td>

                            </tr>

                        {% endfor %}

                    </tbody>

                </table>

            </div>

        </div>
        """

        return render_page(
            "Ученики",
            body,

            rows=rows,
            branches=branches,
            total_count=total_count,
            selected_branch_ids=selected_branch_ids,
            export_url=export_url,
            calculate_age=calculate_age,
        )

    # ========================================================
    # EXCEL
    # ========================================================

    def export_students_xlsx():

        role = g.current_user[
            "role"
        ]

        if role not in (
            "branch",
            "attestation",
            "admin",
        ):
            abort(403)

        query, params = (
            build_students_query()
        )

        rows = get_db().execute(
            query
            + """
                ORDER BY
                    branches.name,
                    students.class_number,
                    students.last_name,
                    students.first_name,
                    students.middle_name
            """,
            params,
        ).fetchall()

        workbook = Workbook()

        worksheet = workbook.active

        worksheet.title = (
            "Ученики"
        )

        # --------------------------------
        # Заголовок
        # --------------------------------

        headers = [
            "№ п/п",
            "ФИО",
            "Филиал",
            "Класс",
            "Возраст",
        ]

        worksheet.append(
            headers
        )

        # Цвета портала:
        # желтый / черный / белый.
        header_fill = PatternFill(
            fill_type="solid",
            fgColor="FFD500",
        )

        header_font = Font(
            bold=True,
            color="000000",
        )

        thin_side = Side(
            style="thin",
            color="D9D9D9",
        )

        cell_border = Border(
            left=thin_side,
            right=thin_side,
            top=thin_side,
            bottom=thin_side,
        )

        for cell in worksheet[1]:

            cell.fill = (
                header_fill
            )

            cell.font = (
                header_font
            )

            cell.alignment = (
                Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )
            )

            cell.border = (
                cell_border
            )

        # --------------------------------
        # Данные
        # --------------------------------

        for index, student in (
            enumerate(
                rows,
                start=1,
            )
        ):

            worksheet.append(
                [
                    index,

                    full_name(
                        student[
                            "last_name"
                        ],
                        student[
                            "first_name"
                        ],
                        student[
                            "middle_name"
                        ],
                    ),

                    student[
                        "branch_name"
                    ],

                    student[
                        "class_number"
                    ],

                    calculate_age(
                        student[
                            "birth_date"
                        ]
                    ),
                ]
            )

        # --------------------------------
        # Оформление
        # --------------------------------

        worksheet.freeze_panes = (
            "A2"
        )

        worksheet.auto_filter.ref = (
            worksheet.dimensions
        )

        worksheet.row_dimensions[
            1
        ].height = 30

        widths = {
            1: 8,
            2: 42,
            3: 32,
            4: 10,
            5: 12,
        }

        for (
            column_number,
            width,
        ) in widths.items():

            worksheet.column_dimensions[
                get_column_letter(
                    column_number
                )
            ].width = width

        for row in (
            worksheet.iter_rows(
                min_row=2,
            )
        ):

            for cell in row:

                cell.border = (
                    cell_border
                )

                cell.alignment = (
                    Alignment(
                        vertical="center",
                        wrap_text=True,
                    )
                )

            # №, класс, возраст —
            # по центру.
            for column in (
                1,
                4,
                5,
            ):

                row[
                    column - 1
                ].alignment = (
                    Alignment(
                        horizontal="center",
                        vertical="center",
                    )
                )

        output = io.BytesIO()

        workbook.save(
            output
        )

        output.seek(0)

        today = (
            datetime.now()
            .strftime(
                "%d-%m-%Y"
            )
        )

        filename = (
            f"Ученики_{today}.xlsx"
        )

        if audit:

            try:
                audit(
                    "students_list_exported",
                    details=(
                        f"rows={len(rows)}; "
                        f"filters="
                        f"{dict(request.args)}"
                    ),
                )
            except Exception:
                pass

        return send_file(
            output,

            as_attachment=True,

            download_name=filename,

            mimetype=(
                "application/"
                "vnd.openxmlformats-"
                "officedocument."
                "spreadsheetml.sheet"
            ),
        )

    # ========================================================
    # ПОДКЛЮЧЕНИЕ
    # ========================================================

    # Старый URL /students уже существует.
    # Просто заменяем его обработчик,
    # поэтому ссылки и меню
    # остаются прежними.

    app.view_functions[
        "students"
    ] = login_required(
        students_v9_7
    )

    # Новый endpoint Excel.
    if (
        "students_export_v9_7"
        not in app.view_functions
    ):

        app.add_url_rule(
            "/students/export.xlsx",

            endpoint=(
                "students_export_v9_7"
            ),

            view_func=login_required(
                export_students_xlsx
            ),

            methods=[
                "GET",
            ],
        )

    print(
        "Students Tools v9.7 "
        "подключен."
    )